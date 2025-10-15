# archivo: caja_operaciones.py
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import datetime
from db_utils import get_connection
try:
    from db_migrations import backup_db
except Exception:
    # best-effort import; backup may not be available in some packaging contexts
    backup_db = None
import sqlite3
import os
import shutil
from theme import (
    COLORS, FONTS, FINANCE_COLORS, format_currency,
    themed_button, apply_theme
)
class DetalleCajaFrame(tk.Frame):
    def __init__(self, parent, caja_id, on_close=None, disable_movimientos=True):
        super().__init__(parent)
        self.caja_id = caja_id
        self.on_close = on_close
        # cache for loaded KPI icons to keep PhotoImage references
        self._kpi_icons = {}
        # Flag to indicate we want to disable ingreso/retiro while this frame is open
        self.disable_movimientos = disable_movimientos
        self.config(bg=COLORS['background'], padx=15, pady=15)

        # Grid configuration (dar más peso al panel izquierdo)
        self.columnconfigure(0, weight=4)  # KPIs y detalles (scrollable)
        self.columnconfigure(1, weight=1)  # Panel de cierre más angosto
        # Asegurar expansión vertical al maximizar
        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        # Panel izquierdo con scroll (KPIs + tablas dentro de un Canvas)
        self._left_container = tk.Frame(self, bg=COLORS['background'])
        self._left_container.grid(row=0, column=0, rowspan=2, sticky='nsew')
        self._left_canvas = tk.Canvas(self._left_container, bg=COLORS['background'], highlightthickness=0)
        self._left_scrollbar = tk.Scrollbar(self._left_container, orient='vertical', command=self._left_canvas.yview)
        self._left_canvas.configure(yscrollcommand=self._left_scrollbar.set)
        self._left_canvas.pack(side='left', fill='both', expand=True)
        self._left_scrollbar.pack(side='right', fill='y')
        # Contenido interno scrollable
        self._left_inner = tk.Frame(self._left_canvas, bg=COLORS['background'])
        self._left_window = self._left_canvas.create_window((0, 0), window=self._left_inner, anchor='nw')

        # Vincular tamaños para scroll
        def _on_left_inner_config(event=None):
            try:
                self._left_canvas.configure(scrollregion=self._left_canvas.bbox('all'))
            except Exception:
                pass
        self._left_inner.bind('<Configure>', _on_left_inner_config)

        def _on_left_container_resize(event=None):
            try:
                self._left_canvas.itemconfig(self._left_window, width=self._left_canvas.winfo_width())
            except Exception:
                pass
        self._left_canvas.bind('<Configure>', _on_left_container_resize)

        # Scroll con rueda del mouse (Windows)
        def _on_mousewheel(event):
            try:
                delta = int(-1*(event.delta/120))
                self._left_canvas.yview_scroll(delta, 'units')
            except Exception:
                pass
        self._left_canvas.bind('<MouseWheel>', _on_mousewheel)

        # Construir paneles dentro del contenedor scrollable (izquierda)
        self._create_kpi_panel(parent=self._left_inner)
        self._create_tables_panel(parent=self._left_inner)
        # Panel derecho (cierre)
        self._create_closure_panel()
        self._load_data()
        # Ask parent (usually main app) to disable menu movimientos while detail is open
        # Menu entries for Ingreso/Retiro were removed from the main menu; nothing to disable here.
        
    def _create_kpi_panel(self, parent=None):
        container = parent if parent is not None else self
        # Panel principal de KPIs
        kpi_frame = tk.Frame(container, bg=COLORS['background'])
        if parent is None:
            kpi_frame.grid(row=0, column=0, sticky="nsew", pady=(0,15))
        else:
            kpi_frame.pack(fill='x', pady=(0, 15))
        # Frame para los KPIs (primera fila)
        self.kpis = tk.Frame(kpi_frame, bg=COLORS['background'])
        self.kpis.pack(fill='x', pady=(0,6))
        # Segunda fila de KPIs para elementos más anchos (ventas totales)
        self.kpis_row2 = tk.Frame(kpi_frame, bg=COLORS['background'])
        self.kpis_row2.pack(fill='x', pady=(6,10))
        
    def create_kpi(self, parent, icon, title, value, bg_color, fg_color):
        # Ajuste: reducir ligeramente la tipografía para KPIs del detalle
        kpi_padx = 12
        kpi_pady = 8
        f = tk.Frame(parent, bg=bg_color, padx=kpi_padx, pady=kpi_pady)
        f.pack(side='left', padx=5)
        try:
            # Título un poco más chico que antes
            title_font = (FONTS['normal'][0], max(FONTS['normal'][1] + 0, 12))
        except Exception:
            title_font = FONTS['normal']
        try:
            # Valor también reducido respecto al anterior
            value_font = (FONTS['bold'][0], max(FONTS['bold'][1] + 2, 14), 'bold')
        except Exception:
            value_font = FONTS['bold']
        # Try to load a PNG icon from img/ when icon is a known key (like 'emitidos', 'anulados', 'ventas_totales', 'ingresos', 'retiros')
        img_label_kwargs = {}
        label_text = f"{icon} {title}"
        try:
            # map of known icon keys to fallback emojis
            fallback = {
                'ingresos': '⬆️',
                'retiros': '⬇️',
                'anulados': '🚫',
                'ventas_totales': '💸',
                'emitidos': '🎟️'
            }
            use_image = False
            img_obj = None
            if isinstance(icon, str) and icon in fallback:
                img_name = f"{icon}.png"
                img_path = os.path.join(os.path.dirname(__file__), 'img', img_name)
                if os.path.exists(img_path):
                    try:
                        # cache PhotoImage to avoid GC
                        if icon not in self._kpi_icons:
                            self._kpi_icons[icon] = tk.PhotoImage(file=img_path)
                        img_obj = self._kpi_icons[icon]
                        use_image = True
                    except Exception:
                        use_image = False
                if not use_image:
                    # fallback to emoji
                    label_text = f"{fallback.get(icon, '')} {title}"
            else:
                # if icon looks like an emoji or text, show it inline
                label_text = f"{icon} {title}"

            if use_image and img_obj is not None:
                tk.Label(f, text=title, image=img_obj, compound='left', bg=bg_color, fg=fg_color, font=title_font).pack()
            else:
                tk.Label(f, text=label_text, bg=bg_color, fg=fg_color, font=title_font).pack()
        except Exception:
            tk.Label(f, text=f"{icon} {title}", bg=bg_color, fg=fg_color, font=title_font).pack()
        val_lbl = tk.Label(f, text=value, bg=bg_color, fg=fg_color, font=value_font)
        val_lbl.pack()
        return f
        
    def _create_tables_panel(self, parent=None):
        container = parent if parent is not None else self
        tables_frame = tk.Frame(container, bg=COLORS['background'])
        if parent is None:
            tables_frame.grid(row=1, column=0, sticky="nsew")
        else:
            tables_frame.pack(fill='x')

        # Button to view/add movimientos (ingresos/retiros)
        btn_mov = themed_button(tables_frame, text="Ver Movimientos (Ingresos/Retiros)", command=self._open_movimientos_window)
        btn_mov.pack(fill='x', pady=(0,8))

        # Ventas por categoría
        cat_frame = tk.LabelFrame(
            tables_frame, text="🗂 Ventas por categoría",
            bg=COLORS['surface'], font=FONTS['bold']
        )
        cat_frame.pack(fill='x', pady=(0, 10))

        from theme import apply_treeview_style
        _style = apply_treeview_style()
        self.cat_tree = ttk.Treeview(
            cat_frame,
            columns=('categoria', 'monto'),
            show='headings',
            height=2,
            style='App.Treeview'
        )
        self.cat_tree.heading('categoria', text='Categoría')
        self.cat_tree.heading('monto', text='Monto')
        self.cat_tree.pack(fill='x', padx=5, pady=5)

        # Productos vendidos
        prod_frame = tk.LabelFrame(
            tables_frame, text="📦 Productos vendidos",
            bg=COLORS['surface'], font=FONTS['bold']
        )
        prod_frame.pack(fill='x')

        self.prod_tree = ttk.Treeview(
            prod_frame,
            columns=('producto', 'cant', 'monto'),
            show='headings',
            height=12,
            style='App.Treeview'
        )
        self.prod_tree.heading('producto', text='Producto')
        self.prod_tree.heading('cant', text='Cant')
        self.prod_tree.heading('monto', text='Monto')
        self.prod_tree.column('cant', width=70, anchor='center')
        self.prod_tree.column('monto', width=100, anchor='e')
        self.prod_tree.pack(fill='x', padx=5, pady=5)
        
    def _create_closure_panel(self):
        closure_frame = tk.LabelFrame(
            self,
            text="Cierre de Caja",
            bg=COLORS['surface'],
            font=(FONTS['bold'][0], max(FONTS['bold'][1] - 1, 11)),
        )
        closure_frame.grid(row=0, column=1, rowspan=2, sticky="nsew", padx=(15, 0))

        # Título/observaciones de apertura (solo lectura)
        label_bigger_font = (FONTS['normal'][0], max(FONTS['normal'][1] - 1, 10))
        tk.Label(
            closure_frame,
            text="📝 Observaciones de apertura:",
            bg=COLORS['surface'],
            font=label_bigger_font,
            fg=COLORS['text'],
        ).pack(anchor='w', pady=(10, 0))
        self.obs_apertura_text = tk.Text(
            closure_frame,
            height=2,
            font=(label_bigger_font[0], max(label_bigger_font[1] - 1, 9)),
            state='disabled',
            bg=COLORS.get('disabled_bg', '#f5f5f5'),
            fg=COLORS.get('disabled_fg', '#666'),
        )
        self.obs_apertura_text.pack(fill='x', pady=(0, 6))

        # Campos de cierre (editables si la caja está abierta)
        self.campos_cierre_frame = tk.Frame(closure_frame, bg=COLORS['surface'])
        self.campos_cierre_frame.pack(fill='x', pady=4)

        tk.Label(
            self.campos_cierre_frame,
            text="💵 Conteo efectivo final en caja:",
            bg=COLORS['surface'],
            font=label_bigger_font,
        ).pack(anchor='w')
        self.conteo_entry = tk.Entry(
            self.campos_cierre_frame,
            font=(label_bigger_font[0], max(label_bigger_font[1] - 1, 9)),
            disabledbackground=COLORS.get('disabled_bg'),
            disabledforeground=COLORS.get('disabled_fg'),
        )
        self.conteo_entry.pack(fill='x', pady=(0, 6))

        tk.Label(
            self.campos_cierre_frame,
            text="🔁 Monto transferencias:",
            bg=COLORS['surface'],
            font=label_bigger_font,
        ).pack(anchor='w')
        self.transf_entry = tk.Entry(
            self.campos_cierre_frame,
            font=(label_bigger_font[0], max(label_bigger_font[1] - 1, 9)),
            disabledbackground=COLORS.get('disabled_bg'),
            disabledforeground=COLORS.get('disabled_fg'),
        )
        self.transf_entry.pack(fill='x', pady=(0, 6))

        tk.Label(
            self.campos_cierre_frame,
            text="👤 Usuario cierre:",
            bg=COLORS['surface'],
            font=label_bigger_font,
        ).pack(anchor='w')
        self.usuario_entry = tk.Entry(
            self.campos_cierre_frame,
            font=(label_bigger_font[0], max(label_bigger_font[1] - 1, 9)),
            disabledbackground=COLORS.get('disabled_bg'),
            disabledforeground=COLORS.get('disabled_fg'),
        )
        self.usuario_entry.pack(fill='x', pady=(0, 6))

        tk.Label(
            self.campos_cierre_frame,
            text="📝 Observaciones de cierre:",
            bg=COLORS['surface'],
            font=label_bigger_font,
        ).pack(anchor='w')
        self.obs_text = tk.Text(
            self.campos_cierre_frame,
            height=2,
            font=(label_bigger_font[0], max(label_bigger_font[1] - 1, 9)),
        )
        self.obs_text.pack(fill='x', pady=(0, 6))

        # Diferencia calculada
        self.diff_label = tk.Label(
            closure_frame,
            text="🧮 Diferencia: $ 0,00",
            font=(FONTS['bold'][0], max(FONTS['bold'][1] - 1, 12), 'bold'),
            bg=COLORS['surface'],
        )
        self.diff_label.pack(pady=6)

        # Botones de acción
        btn_frame = tk.Frame(closure_frame, bg=COLORS['surface'])
        btn_frame.pack(fill='x', pady=6)
        self.btn_cerrar = themed_button(
            btn_frame, text="⚠️ Cerrar Caja", command=self._cerrar_caja,
            bg='#F44336', fg='white'
        )
        self.btn_cerrar.pack(fill='x', pady=(0, 4))
        self.btn_imprimir = themed_button(
            btn_frame, text="🖨️ Imprimir Ticket",
            command=self._imprimir_ticket
        )
        self.btn_imprimir.pack(fill='x', pady=2)
        themed_button(btn_frame, text="📊 Exportar Excel",
                      command=self._exportar_excel).pack(fill='x', pady=2)
    # Export PDF button removed from closure panel; export available from preview

        # Bind para cálculo en vivo
        self.conteo_entry.bind('<KeyRelease>', self._calcular_diferencia)
        self.transf_entry.bind('<KeyRelease>', self._calcular_diferencia)
    
            
    def _calcular_diferencia(self, event=None):
        try:
            def to_float(v):
                try:
                    if v is None or v == '':
                        return 0.0
                    return float(str(v).replace(',', '.'))
                except Exception:
                    return 0.0

            conteo = to_float(self.conteo_entry.get() if hasattr(self, 'conteo_entry') else 0)
            transf = to_float(self.transf_entry.get() if hasattr(self, 'transf_entry') else getattr(self, 'transferencias_final', 0))
            ingresos = to_float(getattr(self, 'ingresos', 0))
            retiros = to_float(getattr(self, 'retiros', 0))
            fondo = to_float(getattr(self, 'fondo_inicial', 0))
            total_ventas = to_float(getattr(self, 'total_ventas', 0))

            # Nueva lógica: real = conteo + transferencias
            real = conteo + transf
            # teor = fondo_inicial + total_ventas + ingresos - retiros
            teor = fondo + total_ventas + ingresos - retiros
            diferencia = real - teor

            color = FINANCE_COLORS['positive_fg'] if diferencia >= 0 else FINANCE_COLORS['negative_fg']
            self.diff_label.config(
                text=f"🧮 Diferencia: {format_currency(diferencia, include_sign=True)}",
                fg=color
            )
            return diferencia
        except Exception:
            self.diff_label.config(
                text="🧮 Diferencia: $ 0,00",
                fg=COLORS['text']
            )
            return 0.0

    def _open_movimientos_window(self):
        """Open a modal window listing caja_movimiento entries for this caja and allow adding new ones.

        This modal centers on screen, uses larger fonts, and defers KPI reload until the
        modal is closed. New movimientos set a flag so the main KPIs are refreshed once.
        """
        try:
            win = tk.Toplevel(self)
            win.title('Movimientos de caja')
            # make modal and centered
            root = self.winfo_toplevel()
            win.transient(root)
            win.grab_set()

            # Make modal larger for easier inspection on tablets/desktops
            w_modal, h_modal = 1000, 700
            sw = win.winfo_screenwidth(); sh = win.winfo_screenheight()
            x = (sw - w_modal) // 2; y = (sh - h_modal) // 2
            win.geometry(f"{w_modal}x{h_modal}+{x}+{y}")

            modal_font = (FONTS['normal'][0], max(FONTS['normal'][1] + 2, 14))
            header_font = (FONTS['bold'][0], max(FONTS['bold'][1] + 2, 16))

            # flag to indicate cambios
            self._movements_changed = False

            top_frame = tk.Frame(win)
            top_frame.pack(fill='x', pady=8)
            tk.Label(top_frame, text=f"Movimientos - Caja: {getattr(self,'codigo_caja', self.caja_id)}", font=header_font).pack(side='left')

            list_frame = tk.Frame(win)
            list_frame.pack(fill='both', expand=True, padx=8, pady=8)

            cols = ('tipo','monto','observacion','creado_ts')
            style = ttk.Style()
            style.configure('Modal.Treeview', font=(FONTS['normal'][0], max(FONTS['normal'][1]+2, 13)), rowheight=28)
            tree = ttk.Treeview(list_frame, columns=cols, show='headings', style='Modal.Treeview')
            tree.heading('tipo', text='Tipo')
            tree.heading('monto', text='Monto')
            tree.heading('observacion', text='Observación')
            tree.heading('creado_ts', text='Fecha')
            tree.column('tipo', width=120, anchor='center')
            tree.column('monto', width=140, anchor='e')
            tree.column('observacion', width=520, anchor='w')
            tree.column('creado_ts', width=260, anchor='center')
            tree.pack(fill='both', expand=True, side='left')

            sb = tk.Scrollbar(list_frame, command=tree.yview)
            sb.pack(side='right', fill='y')
            tree.config(yscrollcommand=sb.set)

            # form to add movement (larger fonts)
            form = tk.Frame(win)
            form.pack(fill='x', padx=8, pady=8)
            tk.Label(form, text='Tipo:', font=modal_font).grid(row=0, column=0, sticky='e')
            tipo_var = tk.StringVar(value='INGRESO')
            tipo_menu = ttk.Combobox(form, textvariable=tipo_var, values=['INGRESO','RETIRO'], width=14, state='readonly')
            tipo_menu.grid(row=0, column=1, sticky='w', padx=6)
            tipo_menu.configure(font=modal_font)
            tk.Label(form, text='Monto:', font=modal_font).grid(row=0, column=2, sticky='e')
            monto_entry = tk.Entry(form, font=modal_font)
            monto_entry.grid(row=0, column=3, sticky='w', padx=6)
            tk.Label(form, text='Observación:', font=modal_font).grid(row=1, column=0, sticky='e')
            obs_entry = tk.Entry(form, width=60, font=modal_font)
            obs_entry.grid(row=1, column=1, columnspan=3, sticky='w', padx=6)

            def _refresh_list():
                for r in tree.get_children():
                    tree.delete(r)
                try:
                    with get_connection() as conn2:
                        cur2 = conn2.cursor()
                        cur2.execute("SELECT tipo, monto, observacion, creado_ts FROM caja_movimiento WHERE caja_id=? ORDER BY creado_ts DESC", (self.caja_id,))
                        for tipo, monto, obs, creado in cur2.fetchall():
                            tree.insert('', 'end', values=(tipo, format_currency(monto), obs or '', creado))
                except Exception:
                    pass

            def _add_movimiento():
                try:
                    tipo = tipo_var.get()
                    monto = float(monto_entry.get().replace(',','.'))
                    obs = obs_entry.get().strip()
                    with get_connection() as conn2:
                        cur2 = conn2.cursor()
                        # Insertar SIEMPRE un nuevo movimiento; triggers mantienen caja_diaria
                        try:
                            cur2.execute(
                                "INSERT INTO caja_movimiento (caja_id, tipo, monto, observacion) VALUES (?, ?, ?, ?)",
                                (self.caja_id, tipo.upper(), monto, obs)
                            )
                        except sqlite3.IntegrityError as e:
                            conn2.rollback()
                            messagebox.showerror(
                                'BD antigua',
                                'Tu base tiene una restricción única en (caja_id, tipo) que impide múltiples ingresos/retiros. Hay que quitarla para continuar.'
                            )
                            return
                        conn2.commit()
                    monto_entry.delete(0, tk.END)
                    obs_entry.delete(0, tk.END)
                    _refresh_list()
                    # mark that movements changed; defer reloading KPIs until modal close
                    try:
                        self._movements_changed = True
                    except Exception:
                        pass
                except Exception as e:
                    messagebox.showerror('Error', f'No se pudo agregar movimiento: {e}')

            btn_add = themed_button(form, text='Agregar movimiento', command=_add_movimiento)
            btn_add.grid(row=2, column=0, columnspan=4, pady=8)
            # If this caja is closed, disable the add button
            try:
                if str(getattr(self, 'estado', '')).lower() == 'cerrada':
                    btn_add.config(state='disabled')
            except Exception:
                pass

            _refresh_list()

            def _on_modal_close():
                try:
                    win.grab_release()
                except Exception:
                    pass
                try:
                    if getattr(self, '_movements_changed', False):
                        try:
                            self._load_data()
                        except Exception:
                            pass
                finally:
                    try:
                        win.destroy()
                    except Exception:
                        pass

            win.protocol('WM_DELETE_WINDOW', _on_modal_close)
            # also block until window closed
            # win.wait_window(win)
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo abrir ventana de movimientos: {e}')
            
    def _load_data(self):
        with get_connection() as conn:
            cursor = conn.cursor()
            # Cargar datos básicos de la caja (usar nombres de columnas para evitar dependencias de índice)
            try:
                cursor.execute("""
                    SELECT cd.id, cd.codigo_caja, cd.fecha, cd.hora_apertura, cd.hora_cierre,
                           cd.usuario_apertura, cd.usuario_cierre, cd.fondo_inicial,
                           cd.observaciones_apertura, cd.obs_cierre, cd.total_ventas,
                           cd.total_efectivo_teorico, cd.conteo_efectivo_final, cd.transferencias_final,
                           cd.ingresos, cd.retiros, cd.diferencia, cd.total_tickets, cd.estado,
                           d.descripcion as nombre_disciplina
                    FROM caja_diaria cd
                    LEFT JOIN disciplinas d ON d.codigo = cd.disciplina
                    WHERE cd.id = ?
                """, (self.caja_id,))
                row = cursor.fetchone()
                if row:
                    cols = [c[0] for c in cursor.description]
                    caja = dict(zip(cols, row))
                else:
                    caja = None
            except sqlite3.Error:
                caja = None
            if not caja:
                messagebox.showerror("Error", "No se encontró la caja especificada")
                if self.on_close:
                    self.on_close()
                return
            # Calcular totales (seguro ante tablas faltantes)
            total_ventas = 0
            try:
                cursor.execute("""
                    SELECT COALESCE(SUM(t.total_ticket),0) as total
                    FROM tickets t 
                    JOIN ventas v ON v.id = t.venta_id
                    WHERE v.caja_id = ? AND t.status != 'Anulado'
                """, (self.caja_id,))
                total_ventas = cursor.fetchone()[0] or 0
            except sqlite3.Error:
                total_ventas = 0

            # Mapear campos de caja a atributos usando nombres de columna
            try:
                self.codigo_caja = caja.get('codigo_caja')
                self.fecha = caja.get('fecha')
                self.hora_apertura = caja.get('hora_apertura')
                self.hora_cierre = caja.get('hora_cierre')
                self.usuario_apertura = caja.get('usuario_apertura')
                self.usuario_cierre = caja.get('usuario_cierre')
                self.fondo_inicial = caja.get('fondo_inicial') or 0
                self.observaciones_apertura = caja.get('observaciones_apertura') or ''
                self.obs_cierre_db = caja.get('obs_cierre') or ''
                # total_ventas puede venir vacío en la tabla
                self.total_ventas = caja.get('total_ventas') or total_ventas or 0
                self.total_teorico = caja.get('total_efectivo_teorico') if caja.get('total_efectivo_teorico') is not None else None
                self.conteo_efectivo_final = caja.get('conteo_efectivo_final') or 0
                self.transferencias_final = caja.get('transferencias_final') or 0
                self.ingresos = caja.get('ingresos') or 0
                self.retiros = caja.get('retiros') or 0
                self.diferencia_db = caja.get('diferencia')
                self.total_tickets = caja.get('total_tickets') or 0
                # contar tickets anulados explícitamente
                try:
                    cursor.execute("SELECT COALESCE(COUNT(*),0) FROM tickets t JOIN ventas v ON v.id = t.venta_id WHERE v.caja_id=? AND t.status='Anulado'", (self.caja_id,))
                    self.tickets_anulados = cursor.fetchone()[0] or 0
                except Exception:
                    self.tickets_anulados = 0
                self.nombre_disciplina = caja.get('nombre_disciplina') or ''
                # estado de la caja
                self.estado = caja.get('estado') or ''
            except Exception:
                # Si algo falla, inicializar valores por seguridad
                self.codigo_caja = None
                self.fecha = self.hora_apertura = self.hora_cierre = self.usuario_apertura = self.usuario_cierre = ''
                self.fondo_inicial = 0
                self.observaciones_apertura = self.obs_cierre_db = ''
                self.total_ventas = total_ventas or 0
                self.total_teorico = getattr(self, 'total_teorico', 0)
                self.conteo_efectivo_final = self.transferencias_final = self.ingresos = self.retiros = 0
                self.diferencia_db = None
                self.total_tickets = 0
                self.nombre_disciplina = ''
                self.estado = ''
            # Si total_teorico no está presente en la fila, estimar con una suma segura
            if getattr(self, 'total_teorico', None) is None:
                try:
                    cursor.execute("""
                        SELECT COALESCE(SUM(t.total_ticket),0)
                        FROM tickets t
                        JOIN ventas v ON v.id = t.venta_id
                        LEFT JOIN metodos_pago mp ON mp.id = v.metodo_pago_id
                        WHERE v.caja_id = ? AND (mp.descripcion='Efectivo' OR mp.descripcion IS NULL) AND t.status != 'Anulado'
                    """, (self.caja_id,))
                    efectivo_ventas = cursor.fetchone()[0] or 0
                except Exception:
                    efectivo_ventas = total_ventas
                # Para evitar depender de posibles valores desactualizados en caja_diaria,
                # obtener ingresos y retiros sumados directamente desde caja_movimiento
                try:
                    cursor.execute(
                        "SELECT COALESCE(SUM(CASE WHEN tipo='INGRESO' THEN monto END),0),\n                                COALESCE(SUM(CASE WHEN tipo='RETIRO' THEN monto END),0)\n                         FROM caja_movimiento WHERE caja_id=?",
                        (self.caja_id,)
                    )
                    _mv = cursor.fetchone() or (0, 0)
                    ingresos_sum = _mv[0] or 0
                    retiros_sum = _mv[1] or 0
                except Exception:
                    ingresos_sum = getattr(self, 'ingresos', 0) or 0
                    retiros_sum = getattr(self, 'retiros', 0) or 0
                # Actualizar atributos para mantener consistencia en el resto de la vista/exports
                try:
                    self.ingresos = ingresos_sum
                    self.retiros = retiros_sum
                except Exception:
                    pass
                # Estimación: fondo inicial + efectivo ventas + ingresos - retiros + transferencias
                self.total_teorico = self.fondo_inicial + efectivo_ventas + ingresos_sum - retiros_sum + self.transferencias_final
            # Poblamos observaciones de apertura
            try:
                self.obs_apertura_text.config(state='normal')
                self.obs_apertura_text.delete('1.0', tk.END)
                self.obs_apertura_text.insert('1.0', getattr(self, 'observaciones_apertura', ''))
                self.obs_apertura_text.config(state='disabled')
            except Exception:
                pass
            # Listar movimientos y sus observaciones en el textbox correspondiente
            # Listar movimientos: almacenar en lista y mostrar solo uno a la vez con navegación
            self._movimientos_list = []
            self._mov_index = 0
            try:
                cursor.execute("SELECT tipo, monto, observacion, creado_ts FROM caja_movimiento WHERE caja_id=? ORDER BY creado_ts", (self.caja_id,))
                for tipo, monto, observacion, creado in cursor.fetchall():
                    text = f"{tipo}: {format_currency(monto)} - {observacion or ''} ({creado})"
                    self._movimientos_list.append(text)
            except Exception:
                self._movimientos_list = []
            try:
                self._mov_index = max(0, len(self._movimientos_list) - 1)
            except Exception:
                # keep internal list empty on error
                self._movimientos_list = []
            # Colocar observacion de cierre en el campo editable
            try:
                self.obs_text.delete('1.0', tk.END)
                if getattr(self, 'obs_cierre_db', ''):
                    self.obs_text.insert('1.0', getattr(self, 'obs_cierre_db', ''))
            except Exception:
                pass
            # Rellenar los campos de cierre con los valores guardados (si existen)
            try:
                # conteo, transferencias, usuario
                try:
                    self.conteo_entry.delete(0, tk.END)
                    self.conteo_entry.insert(0, str(getattr(self, 'conteo_efectivo_final', '') or ''))
                except Exception:
                    pass
                try:
                    self.transf_entry.delete(0, tk.END)
                    self.transf_entry.insert(0, str(getattr(self, 'transferencias_final', '') or ''))
                except Exception:
                    pass
                try:
                    self.usuario_entry.delete(0, tk.END)
                    self.usuario_entry.insert(0, str(getattr(self, 'usuario_cierre', '') or ''))
                except Exception:
                    pass
            except Exception:
                pass
            # Ventas por método de pago
            try:
                cursor.execute("""
                    SELECT mp.descripcion, COUNT(*) as cantidad, SUM(t.total_ticket) as total
                    FROM ventas v
                    JOIN tickets t ON t.venta_id = v.id
                    LEFT JOIN metodos_pago mp ON mp.id = v.metodo_pago_id
                    WHERE v.caja_id = ? AND t.status != 'Anulado'
                    GROUP BY mp.descripcion
                """, (self.caja_id,))
                ventas_por_metodo = cursor.fetchall()
            except sqlite3.Error:
                ventas_por_metodo = []
            # KPIs con valores calculados y consistentes
            # Clear previous KPIs to avoid duplicates when reloading
            try:
                for c in list(self.kpis.winfo_children()):
                    try:
                        c.destroy()
                    except Exception:
                        pass
            except Exception:
                pass
            try:
                for c in list(self.kpis_row2.winfo_children()):
                    try:
                        c.destroy()
                    except Exception:
                        pass
            except Exception:
                pass

            try:
                # Fondo inicial
                self.create_kpi(
                    self.kpis, "💰", "Fondo Inicial",
                    format_currency(getattr(self, 'fondo_inicial', 0)),
                    FINANCE_COLORS.get('transfer_bg', '#eee'),
                    FINANCE_COLORS.get('transfer_fg', '#000')
                )
            except Exception:
                self.create_kpi(self.kpis, "💰", "Fondo Inicial", format_currency(getattr(self, 'fondo_inicial', 0)), '#ffffff', '#000000')
            # Ingresos y Retiros: siempre sumar desde caja_movimiento para reflejar los movimientos reales
            try:
                cursor.execute(
                    "SELECT COALESCE(SUM(CASE WHEN tipo='INGRESO' THEN monto END),0),\n                            COALESCE(SUM(CASE WHEN tipo='RETIRO' THEN monto END),0)\n                     FROM caja_movimiento WHERE caja_id=?",
                    (self.caja_id,)
                )
                mv = cursor.fetchone() or (0, 0)
                ingresos_val = mv[0] or 0
                retiros_val = mv[1] or 0
            except Exception:
                ingresos_val = getattr(self, 'ingresos', 0) or 0
                retiros_val = getattr(self, 'retiros', 0) or 0
            # Mantener atributos sincronizados con los valores mostrados
            try:
                self.ingresos = ingresos_val
                self.retiros = retiros_val
            except Exception:
                pass
            try:
                self.create_kpi(self.kpis, "ingresos", "Ingresos", format_currency(ingresos_val), FINANCE_COLORS.get('positive_bg', '#e6ffed'), FINANCE_COLORS.get('positive_fg', '#0a0'))
                self.create_kpi(self.kpis, "retiros", "Retiros", format_currency(retiros_val), FINANCE_COLORS.get('negative_bg', '#ffecec'), FINANCE_COLORS.get('negative_fg', '#c00'))
            except Exception:
                self.create_kpi(self.kpis, "⬆️", "Ingresos", format_currency(ingresos_val), '#ffffff', '#000000')
                self.create_kpi(self.kpis, "⬇️", "Retiros", format_currency(retiros_val), '#ffffff', '#000000')
            # Ventas por medio de pago sumas (efectivo / transferencias)
            total_efectivo = sum(total for metodo, cant, total in ventas_por_metodo if (metodo or '').lower() == 'efectivo') if ventas_por_metodo else 0
            total_transf = sum(total for metodo, cant, total in ventas_por_metodo if (metodo or '').lower() != 'efectivo') if ventas_por_metodo else 0
            try:
                self.create_kpi(self.kpis_row2, "💵", "Efectivo", format_currency(total_efectivo), FINANCE_COLORS.get('transfer_bg', '#fff'), FINANCE_COLORS.get('transfer_fg', '#000'))
                self.create_kpi(self.kpis_row2, "🔁", "Transferencias", format_currency(total_transf), FINANCE_COLORS.get('transfer_bg', '#fff'), FINANCE_COLORS.get('transfer_fg', '#000'))
            except Exception:
                self.create_kpi(self.kpis_row2, "💵", "Efectivo", format_currency(total_efectivo), '#ffffff', '#000000')
                self.create_kpi(self.kpis_row2, "🔁", "Transferencias", format_currency(total_transf), '#ffffff', '#000000')
            # # KPI: tickets anulados
            # try:
            #     self.create_kpi(self.kpis, "🚫", "Anulados", str(getattr(self, 'tickets_anulados', 0)), COLORS.get('surface'), COLORS.get('text'))
            # except Exception:
            #     try:
            #         self.create_kpi(self.kpis, "🚫", "Anulados", str(getattr(self, 'tickets_anulados', 0)), '#ffffff', '#000')
            #     except Exception:
            #         pass
            # Replace single "Ventas totales" highlight with three KPIs in the second row:
            # Anulados | Emitidos | Ventas totales
            try:
                # compute tickets emitted (status != 'Anulado')
                try:
                    cursor.execute("SELECT COUNT(*) FROM tickets t JOIN ventas v ON v.id=t.venta_id WHERE v.caja_id=? AND t.status != 'Anulado'", (self.caja_id,))
                    tickets_emitidos = cursor.fetchone()[0] or 0
                except Exception:
                    tickets_emitidos = getattr(self, 'total_tickets', 0) or 0

                # Anulados KPI
                try:
                    self.create_kpi(self.kpis, "🚫", "Anulados", str(getattr(self, 'tickets_anulados', 0)), '#ffffff', '#000')
                except Exception:
                    pass

                # Emitidos KPI (simple display)
                try:
                    self.create_kpi(self.kpis, "🎟️", "Emitidos", str(tickets_emitidos), '#ffffff', '#000')
                except Exception:
                    pass

                # Ventas totales KPI (destacado a la derecha)
                try:
                    self.create_kpi(self.kpis_row2, "📊", "Ventas totales", format_currency(getattr(self, 'total_ventas', total_ventas)), FINANCE_COLORS.get('total_sales_bg', COLORS.get('surface', '#f5f5f5')), FINANCE_COLORS.get('total_sales_fg', COLORS.get('text', '#000')))
                except Exception:
                    pass
            except Exception:
                pass
            # Cargar tablas: categorías
            try:
                cursor.execute("""
                    SELECT c.descripcion, SUM(t.total_ticket) as total
                    FROM tickets t
                    JOIN ventas v ON v.id = t.venta_id
                    LEFT JOIN Categoria_Producto c ON c.id = t.categoria_id
                    WHERE v.caja_id = ? AND t.status != 'Anulado'
                    GROUP BY c.descripcion
                    ORDER BY total DESC
                """, (self.caja_id,))
                rows = cursor.fetchall()
            except sqlite3.Error:
                rows = []
            for cat in self.cat_tree.get_children():
                self.cat_tree.delete(cat)
            for row in rows:
                cat = row[0] or 'Sin categoría'
                total = row[1] or 0
                self.cat_tree.insert('', 'end', values=(cat, format_currency(total)))
            # Productos vendidos
            try:
                cursor.execute("""
                    SELECT p.nombre, SUM(vi.cantidad) as cant, 
                           SUM(vi.cantidad * vi.precio_unitario) as total
                    FROM venta_items vi
                    JOIN tickets t ON t.id = vi.ticket_id
                    JOIN ventas v ON v.id = t.venta_id
                    JOIN products p ON p.id = vi.producto_id
                    WHERE v.caja_id = ? AND t.status != 'Anulado'
                    GROUP BY p.nombre
                    ORDER BY total DESC
                """, (self.caja_id,))
                prod_rows = cursor.fetchall()
            except sqlite3.Error:
                prod_rows = []
            for prod in self.prod_tree.get_children():
                self.prod_tree.delete(prod)
            for row in prod_rows:
                try:
                    self.prod_tree.insert('', 'end', values=(row[0], int(row[1]), format_currency(row[2] or 0)))
                except Exception:
                    pass
            # Obtener estado actual (cerrada/abierta) desde la fila ya leída
            esta_cerrada = (getattr(self, 'estado', '') == 'cerrada')
            # Ajustar estado de campos de cierre según si la caja está cerrada
            if esta_cerrada:
                # Deshabilitar campos editables del cierre
                try:
                    self.conteo_entry.config(state='disabled')
                    self.transf_entry.config(state='disabled')
                    self.usuario_entry.config(state='disabled')
                    # Observaciones de cierre no editables (gris)
                    self.obs_text.config(state='disabled', bg=COLORS.get('disabled_bg', '#f5f5f5'))
                except Exception:
                    pass
                # Ocultar o deshabilitar botón de cerrar y mostrar/habilitar imprimir
                try:
                    # preferimos deshabilitar botón de cerrar si existe
                    self.btn_cerrar.config(state='disabled')
                except Exception:
                    try:
                        self.btn_cerrar.pack_forget()
                    except Exception:
                        pass
                try:
                    self.btn_imprimir.config(state='normal')
                    self.btn_imprimir.pack(fill='x', pady=2)
                except Exception:
                    try:
                        self.btn_imprimir.pack(fill='x', pady=2)
                    except Exception:
                        pass
            # Mostrar diferencia almacenada si la caja está cerrada; si está abierta, calcular en vivo
            if esta_cerrada:
                # Mostrar valor guardado en la base (self.diferencia_db)
                diff_val = getattr(self, 'diferencia_db', None)
                if diff_val is None:
                    diff_val = 0.0
                color = FINANCE_COLORS['positive_fg'] if float(diff_val) >= 0 else FINANCE_COLORS['negative_fg']
                self.diff_label.config(text=f"🧮 Diferencia: {format_currency(diff_val, include_sign=True)}", fg=color)
            else:
                # Caja abierta: campos editables y botón de cerrar activo; mantener cálculo en vivo
                try:
                    self.conteo_entry.config(state='normal')
                    self.transf_entry.config(state='normal')
                    self.usuario_entry.config(state='normal')
                    self.obs_text.config(state='normal')
                except Exception:
                    pass
                try:
                    self._calcular_diferencia()
                except Exception:
                    pass
                try:
                    self.btn_cerrar.config(state='normal')
                    self.btn_cerrar.pack(fill='x', pady=(0,5))
                except Exception:
                    try:
                        self.btn_cerrar.pack(fill='x', pady=(0,5))
                    except Exception:
                        pass
                try:
                    # ocultar o desactivar imprimir si la caja no está cerrada
                    self.btn_imprimir.config(state='disabled')
                    self.btn_imprimir.pack_forget()
                except Exception:
                    try:
                        self.btn_imprimir.pack_forget()
                    except Exception:
                        pass
            # Guardar total teórico (usar columna total_efectivo_teorico si existe)
            try:
                self.total_teorico = caja[14] or total_ventas
            except Exception:
                self.total_teorico = total_ventas
            
    def _cerrar_caja(self):
        try:
            conteo = float(self.conteo_entry.get().replace(',', '.') or 0)
            transf = float(self.transf_entry.get().replace(',', '.') or 0)
            usuario = self.usuario_entry.get().strip()
            obs = self.obs_text.get("1.0", tk.END).strip()
            
            if not usuario:
                messagebox.showwarning("Validación", "Debe ingresar el usuario de cierre")
                return
                
            if messagebox.askyesno("Confirmar", 
                                 "¿Está seguro de cerrar la caja? Esta acción no se puede deshacer."):
                with get_connection() as conn:
                    cursor = conn.cursor()
                    now = datetime.datetime.now()
                    # Calcular diferencia con la nueva fórmula:
                    ingresos = float(getattr(self, 'ingresos', 0) or 0)
                    retiros = float(getattr(self, 'retiros', 0) or 0)
                    fondo = float(getattr(self, 'fondo_inicial', 0) or 0)
                    total_ventas = float(getattr(self, 'total_ventas', 0) or 0)
                    # real = conteo + transferencias
                    real = conteo + transf
                    # teor = fondo_inicial + total_ventas + ingresos - retiros
                    teorico = fondo + total_ventas + ingresos - retiros
                    diferencia = real - teorico
                    cursor.execute("""
                        UPDATE caja_diaria 
                        SET estado = 'cerrada',
                            hora_cierre = ?,
                            usuario_cierre = ?,
                            conteo_efectivo_final = ?,
                            transferencias_final = ?,
                            diferencia = ?,
                            obs_cierre = ?
                        WHERE id = ?
                    """, (
                        now.strftime('%H:%M:%S'),
                        usuario,
                        conteo,
                        transf,
                        diferencia,
                        obs,
                        self.caja_id
                    ))
                    conn.commit()
                    # Create a timestamped backup after successful close. Use best-effort call.
                    try:
                        if callable(backup_db):
                            # Run backup and inform the user. Use best-effort and log result to AppData
                            try:
                                path = backup_db()
                                try:
                                    # write a small log entry for auditing
                                    from utils_paths import appdata_dir
                                    log_dir = appdata_dir()
                                    log_path = os.path.join(log_dir, 'backup_logs.txt')
                                    with open(log_path, 'a', encoding='utf-8') as lf:
                                        lf.write(f"{datetime.datetime.now().isoformat()} - backup created: {path}\n")
                                except Exception:
                                    pass
                                # Inform the user that backup was created
                                try:
                                    messagebox.showinfo('Backup', f'Backup creado: {path}')
                                except Exception:
                                    pass
                            except Exception as _e:
                                try:
                                    from utils_paths import appdata_dir
                                    log_dir = appdata_dir()
                                    log_path = os.path.join(log_dir, 'backup_logs.txt')
                                    with open(log_path, 'a', encoding='utf-8') as lf:
                                        lf.write(f"{datetime.datetime.now().isoformat()} - backup failed: {_e}\n")
                                except Exception:
                                    pass
                                try:
                                    messagebox.showwarning('Backup', f'No se pudo crear backup: {_e}')
                                except Exception:
                                    pass
                        else:
                            # backup function not available: attempt inline backup using DB_PATH
                            try:
                                from utils_paths import DB_PATH, appdata_dir
                                ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                                bdir = os.path.join(appdata_dir(), 'backup')
                                os.makedirs(bdir, exist_ok=True)
                                dst = os.path.join(bdir, f'barcancha_{ts}.db')
                                try:
                                    src_conn = sqlite3.connect(DB_PATH)
                                    dest_conn = sqlite3.connect(dst)
                                    with dest_conn:
                                        src_conn.backup(dest_conn)
                                    try:
                                        dest_conn.close()
                                    except Exception:
                                        pass
                                    try:
                                        src_conn.close()
                                    except Exception:
                                        pass
                                    # log + notify
                                    try:
                                        log_path = os.path.join(appdata_dir(), 'backup_logs.txt')
                                        with open(log_path, 'a', encoding='utf-8') as lf:
                                            lf.write(f"{datetime.datetime.now().isoformat()} - inline backup created: {dst}\n")
                                    except Exception:
                                        pass
                                    try:
                                        messagebox.showinfo('Backup', f'Backup creado: {dst}')
                                    except Exception:
                                        pass
                                except Exception as _e2:
                                    # fallback to file copy
                                    try:
                                        shutil.copy2(DB_PATH, dst)
                                        try:
                                            log_path = os.path.join(appdata_dir(), 'backup_logs.txt')
                                            with open(log_path, 'a', encoding='utf-8') as lf:
                                                lf.write(f"{datetime.datetime.now().isoformat()} - inline backup copied: {dst}\n")
                                        except Exception:
                                            pass
                                        try:
                                            messagebox.showinfo('Backup', f'Backup creado (copiado): {dst}')
                                        except Exception:
                                            pass
                                    except Exception as _e3:
                                        try:
                                            log_path = os.path.join(appdata_dir(), 'backup_logs.txt')
                                            with open(log_path, 'a', encoding='utf-8') as lf:
                                                lf.write(f"{datetime.datetime.now().isoformat()} - inline backup failed: {_e3}\n")
                                        except Exception:
                                            pass
                                        try:
                                            messagebox.showwarning('Backup', f'No se pudo crear backup local: {_e3}')
                                        except Exception:
                                            pass
                            except Exception:
                                # if even utils_paths import fails, give up silently but log if possible
                                try:
                                    from utils_paths import appdata_dir
                                    log_path = os.path.join(appdata_dir(), 'backup_logs.txt')
                                    with open(log_path, 'a', encoding='utf-8') as lf:
                                        lf.write(f"{datetime.datetime.now().isoformat()} - inline backup fallback failed.\n")
                                except Exception:
                                    pass
                    except Exception:
                        pass
                    messagebox.showinfo("Éxito", "Caja cerrada correctamente")
                    # Notify caller that the caja was actually closed
                    if self.on_close:
                        try:
                            self.on_close(True)
                        except TypeError:
                            # backward compatibility
                            self.on_close()
        
        except ValueError:
            messagebox.showerror("Error", 
                               "Los montos deben ser números válidos")

    # Navigation helpers for movimientos list shown in the closure panel
    def _show_current_movimiento(self):
        # Navigation helper removed: we no longer display movimientos inline in the closure panel.
        # Keeping method as no-op for compatibility.
        try:
            return
        except Exception:
            pass

    def _mov_prev(self):
        try:
            if not getattr(self, '_movimientos_list', None):
                return
            self._mov_index = max(0, self._mov_index - 1)
            self._show_current_movimiento()
        except Exception:
            pass

    def _mov_next(self):
        try:
            if not getattr(self, '_movimientos_list', None):
                return
            self._mov_index = min(len(self._movimientos_list)-1, self._mov_index + 1)
            self._show_current_movimiento()
        except Exception:
            pass
    
    def _exportar_excel(self):
        try:
            # Nombre por defecto usando el codigo legible de la caja si está disponible
            default_code = getattr(self, 'codigo_caja', None) or self.caja_id
            default_name = f"{default_code}_CierreCaja.xlsx"
            filename = filedialog.asksaveasfilename(
                initialfile=default_name,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Guardar reporte de cierre de caja"
            )
            if not filename:
                return
            # Intentar usar openpyxl si está disponible
            try:
                import openpyxl
                from openpyxl.utils import get_column_letter
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "CierreCaja"
                # Cabecera con columnas solicitadas
                columns = [
                    'Codigo Caja', 'Fecha', 'Hora Apertura', 'Hora Cierre',
                    'Usuario Apertura', 'Usuario Cierre', 'Disciplina',
                    'Fondo inicial', 'Total ventas', 'Total efectivo teorico',
                    'Conteo efectivo final', 'Transferencias final', 'Ingresos', 'Retiros',
                    'Diferencia', 'Total tickets', 'Observacion apertura', 'Observacion cierre',
                    'Movimientos', 'Items vendidos'
                ]
                ws.append(columns)
                # Preparar valores (leer desde atributos / widgets con fallbacks)
                # Use internal movimientos list to build export string
                try:
                    movimientos_joined = ' | '.join(self._movimientos_list)
                except Exception:
                    movimientos_joined = ''
                # Construir cadena con items vendidos: Producto (Cant x subtotal)
                try:
                    with get_connection() as conn2:
                        cur2 = conn2.cursor()
                        cur2.execute("""
                            SELECT p.nombre, SUM(vi.cantidad) as cant, SUM(vi.cantidad * vi.precio_unitario) as total
                            FROM venta_items vi
                            JOIN tickets t ON t.id = vi.ticket_id
                            JOIN ventas v ON v.id = t.venta_id
                            JOIN products p ON p.id = vi.producto_id
                            WHERE v.caja_id = ?
                            GROUP BY p.nombre
                        """, (self.caja_id,))
                        items_rows = cur2.fetchall()
                    items_joined = ' | '.join(f"{r[0]} ({int(r[1])} x {format_currency(r[2])})" for r in items_rows)
                except Exception:
                    items_joined = ''
                # compute ingresos/retiros for export (prefer attributes; fall back to summing caja_movimiento)
                try:
                    ingresos_val = getattr(self, 'ingresos', None)
                    retiros_val = getattr(self, 'retiros', None)
                    if ingresos_val is None or retiros_val is None:
                        try:
                            cur2.execute("SELECT COALESCE(SUM(CASE WHEN tipo='INGRESO' THEN monto ELSE 0 END),0), COALESCE(SUM(CASE WHEN tipo='RETIRO' THEN monto ELSE 0 END),0) FROM caja_movimiento WHERE caja_id=?", (self.caja_id,))
                            mv = cur2.fetchone()
                            if mv:
                                if ingresos_val is None:
                                    ingresos_val = mv[0] or 0
                                if retiros_val is None:
                                    retiros_val = mv[1] or 0
                        except Exception:
                            ingresos_val = ingresos_val if ingresos_val is not None else 0
                            retiros_val = retiros_val if retiros_val is not None else 0
                except Exception:
                    ingresos_val = getattr(self, 'ingresos', 0) or 0
                    retiros_val = getattr(self, 'retiros', 0) or 0

                row = [
                    getattr(self, 'codigo_caja', self.caja_id),
                    getattr(self, 'fecha', ''),
                    getattr(self, 'hora_apertura', ''),
                    getattr(self, 'hora_cierre', ''),
                    getattr(self, 'usuario_apertura', ''),
                    getattr(self, 'usuario_cierre', ''),
                    getattr(self, 'nombre_disciplina', ''),
                    getattr(self, 'fondo_inicial', self.conteo_entry.get()),
                    getattr(self, 'total_ventas', ''),
                    getattr(self, 'total_teorico', ''),
                    getattr(self, 'conteo_efectivo_final', self.conteo_entry.get()),
                    getattr(self, 'transferencias_final', self.transf_entry.get()),
                    ingresos_val,
                    retiros_val,
                    getattr(self, 'diferencia_db', self.diff_label.cget('text')),
                    getattr(self, 'total_tickets', ''),
                    getattr(self, 'observaciones_apertura', ''),
                    getattr(self, 'obs_cierre_db', ''),
                    movimientos_joined,
                    items_joined
                ]
                ws.append(row)
                # Ajustar anchos
                for i in range(1, len(columns) + 1):
                    ws.column_dimensions[get_column_letter(i)].width = 25
                wb.save(filename)
            except Exception:
                # fallback simple CSV-like TXT
                with open(filename, 'w', encoding='utf-8') as f:
                    cols = ['Codigo Caja','Fecha','Hora Apertura','Hora Cierre','Usuario Apertura','Usuario Cierre','Disciplina','Fondo inicial','Total ventas','Total efectivo teorico','Conteo efectivo final','Transferencias final','Ingresos','Retiros','Diferencia','Total tickets','Observacion apertura','Observacion cierre','Movimientos','Items vendidos']
                    f.write(';'.join(cols) + '\n')
                    try:
                        mov = ' | '.join(self._movimientos_list)
                    except Exception:
                        mov = ''
                    try:
                        items_joined_val = items_joined
                    except Exception:
                        items_joined_val = ''
                    # compute ingresos/retiros for fallback export
                    try:
                        ingresos_val = getattr(self, 'ingresos', None)
                        retiros_val = getattr(self, 'retiros', None)
                        if ingresos_val is None or retiros_val is None:
                            try:
                                with get_connection() as _conn_tmp:
                                    _cur_tmp = _conn_tmp.cursor()
                                    _cur_tmp.execute("SELECT COALESCE(SUM(CASE WHEN tipo='INGRESO' THEN monto ELSE 0 END),0), COALESCE(SUM(CASE WHEN tipo='RETIRO' THEN monto ELSE 0 END),0) FROM caja_movimiento WHERE caja_id=?", (self.caja_id,))
                                    _mv = _cur_tmp.fetchone()
                                    if _mv:
                                        if ingresos_val is None:
                                            ingresos_val = _mv[0] or 0
                                        if retiros_val is None:
                                            retiros_val = _mv[1] or 0
                            except Exception:
                                ingresos_val = ingresos_val if ingresos_val is not None else 0
                                retiros_val = retiros_val if retiros_val is not None else 0
                    except Exception:
                        ingresos_val = getattr(self, 'ingresos', 0) or 0
                        retiros_val = getattr(self, 'retiros', 0) or 0

                    values = [
                        str(getattr(self, 'codigo_caja', self.caja_id)),
                        str(getattr(self, 'fecha', '')),
                        str(getattr(self, 'hora_apertura', '')),
                        str(getattr(self, 'hora_cierre', '')),
                        str(getattr(self, 'usuario_apertura', '')),
                        str(getattr(self, 'usuario_cierre', '')),
                        str(getattr(self, 'nombre_disciplina', '')),
                        str(getattr(self, 'fondo_inicial', self.conteo_entry.get())),
                        str(getattr(self, 'total_ventas', '')),
                        str(getattr(self, 'total_teorico', '')),
                        str(getattr(self, 'conteo_efectivo_final', self.conteo_entry.get())),
                        str(getattr(self, 'transferencias_final', self.transf_entry.get())),
                        str(ingresos_val),
                        str(retiros_val),
                        str(getattr(self, 'diferencia_db', self.diff_label.cget('text'))),
                        str(getattr(self, 'total_tickets', '')),
                        str(getattr(self, 'observaciones_apertura', '')),
                        str(getattr(self, 'obs_cierre_db', '')),
                        str(mov),
                        str(items_joined_val)
                    ]
                    f.write(';'.join(values) + '\n')
            # Abrir archivo automáticamente
            try:
                import os, subprocess, sys
                if sys.platform.startswith('win'):
                    os.startfile(filename)
                else:
                    subprocess.Popen(['xdg-open', filename])
            except Exception:
                pass
            messagebox.showinfo("Éxito", f"Archivo guardado: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {str(e)}")
    
    def _exportar_pdf(self):
        try:
            default_code = getattr(self, 'codigo_caja', None) or self.caja_id
            default_name = f"{default_code}_CierreCaja.pdf"
            filename = filedialog.asksaveasfilename(
                initialfile=default_name,
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Guardar reporte de cierre de caja"
            )
            if not filename:
                return
            # Intentar usar reportlab
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas
                c = canvas.Canvas(filename, pagesize=A4)
                y = 800
                c.setFont("Helvetica-Bold", 14)
                c.drawCentredString(300, y, "CIERRE DE CAJA")
                y -= 40
                c.setFont("Helvetica", 10)
                c.drawString(50, y, f"Codigo Caja: {getattr(self, 'codigo_caja', self.caja_id)}")
                y -= 20
                c.drawString(50, y, f"Fondo inicial: {getattr(self, 'fondo_inicial', self.conteo_entry.get())}")
                y -= 20
                c.drawString(50, y, f"Transferencias: {getattr(self, 'transferencias_final', self.transf_entry.get())}")
                y -= 20
                # Mostrar Ingresos y Retiros
                try:
                    _ing = getattr(self, 'ingresos', None)
                    _ret = getattr(self, 'retiros', None)
                    if _ing is None or _ret is None:
                        try:
                            with get_connection() as _conn_tmp:
                                _cur_tmp = _conn_tmp.cursor()
                                _cur_tmp.execute("SELECT COALESCE(SUM(CASE WHEN tipo='INGRESO' THEN monto ELSE 0 END),0), COALESCE(SUM(CASE WHEN tipo='RETIRO' THEN monto ELSE 0 END),0) FROM caja_movimiento WHERE caja_id=?", (self.caja_id,))
                                _m = _cur_tmp.fetchone()
                                if _m:
                                    if _ing is None:
                                        _ing = _m[0] or 0
                                    if _ret is None:
                                        _ret = _m[1] or 0
                        except Exception:
                            _ing = _ing if _ing is not None else 0
                            _ret = _ret if _ret is not None else 0
                except Exception:
                    _ing = getattr(self, 'ingresos', 0) or 0
                    _ret = getattr(self, 'retiros', 0) or 0
                try:
                    c.drawString(50, y, f"Ingresos: {format_currency(_ing)}")
                    y -= 20
                    c.drawString(50, y, f"Retiros: {format_currency(-abs(float(_ret)))}")
                    y -= 20
                except Exception:
                    pass
                # Mostrar diferencia con signo (usar valor numérico si es posible)
                try:
                    dlab = getattr(self, 'diferencia_db', None)
                    if dlab is None:
                        # intentar extraer del label
                        lbl = self.diff_label.cget('text')
                        dlab = lbl.split()[-1]
                    dnum = float(str(dlab).replace('$','').replace(',', '.'))
                except Exception:
                    dnum = 0.0
                c.drawString(50, y, f"Diferencia: {format_currency(dnum, include_sign=True)}")
                y -= 20
                try:
                    c.drawString(50, y, f"Tickets anulados: {getattr(self, 'tickets_anulados', 0)}")
                    y -= 20
                except Exception:
                    pass
                # incluir items vendidos
                try:
                    with get_connection() as conn3:
                        ccur = conn3.cursor()
                        ccur.execute("""
                            SELECT p.nombre, SUM(vi.cantidad) as cant, SUM(vi.cantidad * vi.precio_unitario) as total
                            FROM venta_items vi
                            JOIN tickets t ON t.id = vi.ticket_id
                            JOIN ventas v ON v.id = t.venta_id
                            JOIN products p ON p.id = vi.producto_id
                            WHERE v.caja_id = ?
                            GROUP BY p.nombre
                        """, (self.caja_id,))
                        items_rows_pdf = ccur.fetchall()
                    if items_rows_pdf:
                        y -= 10
                        c.setFont("Helvetica-Bold", 12)
                        c.drawString(50, y, "Items vendidos:")
                        y -= 18
                        c.setFont("Helvetica", 10)
                        for r in items_rows_pdf:
                            line = f"({r[0]} x {int(r[1])}) = {format_currency(r[2])}"
                            c.drawString(60, y, line)
                            y -= 14
                            if y < 60:
                                c.showPage()
                                y = 800
                except Exception:
                    pass
                c.save()
            except Exception:
                # fallback: write plain text
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write('CIERRE DE CAJA\n')
                    f.write(f"Codigo Caja: {getattr(self, 'codigo_caja', self.caja_id)}\n")
                    f.write(f"Fondo inicial: {getattr(self, 'fondo_inicial', self.conteo_entry.get())}\n")
                    f.write(f"Transferencias: {getattr(self, 'transferencias_final', self.transf_entry.get())}\n")
                    # Incluir Ingresos y Retiros en texto
                    try:
                        _ing = getattr(self, 'ingresos', None)
                        _ret = getattr(self, 'retiros', None)
                        if _ing is None or _ret is None:
                            with get_connection() as _c:
                                _cur = _c.cursor()
                                _cur.execute("SELECT COALESCE(SUM(CASE WHEN tipo='INGRESO' THEN monto ELSE 0 END),0), COALESCE(SUM(CASE WHEN tipo='RETIRO' THEN monto ELSE 0 END),0) FROM caja_movimiento WHERE caja_id=?", (self.caja_id,))
                                _r = _cur.fetchone()
                                if _r:
                                    if _ing is None:
                                        _ing = _r[0] or 0
                                    if _ret is None:
                                        _ret = _r[1] or 0
                    except Exception:
                        _ing = getattr(self, 'ingresos', 0) or 0
                        _ret = getattr(self, 'retiros', 0) or 0
                    try:
                        f.write(f"Ingresos: {_ing}\n")
                        f.write(f"Retiros: {-abs(float(_ret))}\n")
                    except Exception:
                        f.write(f"Ingresos: {getattr(self, 'ingresos', 0)}\n")
                        f.write(f"Retiros: {-abs(getattr(self, 'retiros', 0) or 0)}\n")
                    # intentar escribir diferencia con signo
                    try:
                        dlab2 = getattr(self, 'diferencia_db', None)
                        if dlab2 is None:
                            lbl2 = self.diff_label.cget('text')
                            dlab2 = lbl2.split()[-1]
                        dnum2 = float(str(dlab2).replace('$','').replace(',', '.'))
                    except Exception:
                        dnum2 = 0.0
                    f.write(f"Diferencia: {format_currency(dnum2, include_sign=True)}\n")
                    try:
                        f.write(f"Tickets anulados: {getattr(self, 'tickets_anulados', 0)}\n")
                    except Exception:
                        pass
                    # incluir items vendidos en texto
                    try:
                        with get_connection() as conn4:
                            tcur = conn4.cursor()
                            tcur.execute("""
                                SELECT p.nombre, SUM(vi.cantidad) as cant, SUM(vi.cantidad * vi.precio_unitario) as total
                                FROM venta_items vi
                                JOIN tickets t ON t.id = vi.ticket_id
                                JOIN ventas v ON v.id = t.venta_id
                                JOIN products p ON p.id = vi.producto_id
                                WHERE v.caja_id = ?
                                GROUP BY p.nombre
                            """, (self.caja_id,))
                            trows = tcur.fetchall()
                            if trows:
                                f.write('\nItems vendidos:\n')
                                for r in trows:
                                    f.write(f"({r[0]} x {int(r[1])}) = {format_currency(r[2])}\n")
                    except Exception:
                        pass
            # abrir automaticamente
            try:
                import os, sys, subprocess
                if sys.platform.startswith('win'):
                    os.startfile(filename)
                else:
                    subprocess.Popen(['xdg-open', filename])
            except Exception:
                pass
            messagebox.showinfo("Éxito", f"Archivo guardado: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {str(e)}")
            
    def _imprimir_ticket(self):
        """Imprime el ticket de cierre de caja"""
        try:
            with get_connection() as conn:
                cursor = conn.cursor()
                # Build ticket using attributes already loaded in the frame (safe fallbacks)
                ticket = []
                ticket.append("=" * 40)
                ticket.append("CIERRE DE CAJA".center(40))
                ticket.append("=" * 40)
                ticket.append(f"Codigo caja: {getattr(self, 'codigo_caja', '')}")
                ticket.append(f"Fecha apertura: {getattr(self, 'fecha', '')} {getattr(self, 'hora_apertura', '')}")
                ticket.append(f"Usuario apertura: {getattr(self, 'usuario_apertura', '')}")
                ticket.append(f"Disciplina: {getattr(self, 'nombre_disciplina', '')}")
                ticket.append(f"Fecha cierre: {getattr(self, 'fecha', '')} {getattr(self, 'hora_cierre', '')}")
                ticket.append(f"Usuario cierre: {getattr(self, 'usuario_cierre', '')}")
                ticket.append("-" * 40)
                
                # Totales por método de pago
                cursor.execute("""
                    SELECT mp.descripcion, SUM(t.total_ticket) as total
                    FROM ventas v
                    JOIN tickets t ON t.venta_id = v.id
                    LEFT JOIN metodos_pago mp ON mp.id = v.metodo_pago_id
                    WHERE v.caja_id = ? AND t.status != 'Anulado'
                    GROUP BY mp.descripcion
                """, (self.caja_id,))
                
                ticket.append("TOTALES POR MEDIO DE PAGO")
                ticket.append("-" * 40)
                total_general = 0
                for metodo, total in cursor.fetchall():
                    ticket.append(f"{metodo}: {format_currency(total)}")
                    total_general += total
                
                ticket.append("-" * 40)
                ticket.append(f"TOTAL: {format_currency(total_general)}")
                ticket.append("-" * 40)
                
                # Información de cierre (usar atributos para robustez)
                try:
                    ticket.append(f"Fondo inicial: {format_currency(getattr(self, 'fondo_inicial', 0))}")
                    ticket.append(f"Conteo final: {format_currency(getattr(self, 'conteo_efectivo_final', getattr(self, 'conteo_entry', '') or 0))}")
                    ticket.append(f"Transferencias: {format_currency(getattr(self, 'transferencias_final', 0))}")
                    # Ingresos / Retiros (mostrar retiros como negativo)
                    try:
                        ingresos_val = getattr(self, 'ingresos', None)
                        retiros_val = getattr(self, 'retiros', None)
                        if ingresos_val is None or retiros_val is None:
                            cursor.execute("SELECT COALESCE(SUM(CASE WHEN tipo='INGRESO' THEN monto ELSE 0 END),0), COALESCE(SUM(CASE WHEN tipo='RETIRO' THEN monto ELSE 0 END),0) FROM caja_movimiento WHERE caja_id=?", (self.caja_id,))
                            _mv = cursor.fetchone()
                            if _mv:
                                if ingresos_val is None:
                                    ingresos_val = _mv[0] or 0
                                if retiros_val is None:
                                    retiros_val = _mv[1] or 0
                    except Exception:
                        ingresos_val = getattr(self, 'ingresos', 0) or 0
                        retiros_val = getattr(self, 'retiros', 0) or 0
                    try:
                        ticket.append(f"Ingresos: {format_currency(ingresos_val)}")
                        ticket.append(f"Retiros: {format_currency(-abs(float(retiros_val)))}")
                    except Exception:
                        ticket.append(f"Ingresos: {format_currency(getattr(self, 'ingresos', 0))}")
                        ticket.append(f"Retiros: {format_currency(-abs(getattr(self, 'retiros', 0) or 0))}")
                    # preferir diferencia almacenada si existe
                    diff_val = getattr(self, 'diferencia_db', None)
                    if diff_val is None:
                        # intentar parsear del label (fallback)
                        try:
                            # label tiene formato '🧮 Diferencia: $ 0,00' o similar
                            label = self.diff_label.cget('text')
                            # extraer último token
                            diff_val = label.split()[-1]
                        except Exception:
                            diff_val = 0
                    # Mostrar diferencia con signo
                    try:
                        dnum = float(diff_val)
                    except Exception:
                        try:
                            # parsear texto con moneda
                            dnum = float(str(diff_val).replace('$', '').replace(',', '.'))
                        except Exception:
                            dnum = 0.0
                    ticket.append(f"Diferencia: {format_currency(dnum, include_sign=True)}")
                except Exception:
                    pass
                # Tickets anulados (desde atributo calculado)
                ticket.append(f"Tickets anulados: {getattr(self, 'tickets_anulados', 0)}")
                ticket.append("=" * 40)
                
                # Items vendidos: Producto, Cantidad, Monto
                try:
                    cursor.execute("""
                        SELECT p.nombre, SUM(vi.cantidad) as cant, SUM(vi.cantidad * vi.precio_unitario) as total
                        FROM venta_items vi
                        JOIN tickets t ON t.id = vi.ticket_id
                        JOIN ventas v ON v.id = t.venta_id
                        JOIN products p ON p.id = vi.producto_id
                        WHERE v.caja_id = ?
                        GROUP BY p.nombre
                        ORDER BY total DESC
                    """, (self.caja_id,))
                    items = cursor.fetchall()
                    if items:
                        ticket.append("ITEMS VENDIDOS:")
                        ticket.append("(Producto x Cant) = Monto Total")
                        ticket.append("-" * 40)
                        for nombre, cant, total in items:
                            ticket.append(f"({nombre} x {int(cant)}) = {format_currency(total)}")
                except Exception:
                    pass

                # Mostrar vista previa en ventana con acciones: Imprimir / Exportar a PDF
                preview = "\n".join(ticket)
                try:
                    win = tk.Toplevel(self)
                    win.title('Vista previa ticket')
                    txt = tk.Text(win, width=60, height=30, wrap='none')
                    txt.insert('1.0', preview)
                    txt.config(state='disabled')
                    txt.pack(fill='both', expand=True)
                    btn_frame = tk.Frame(win)
                    btn_frame.pack(fill='x', pady=6)

                    # (Se omite control de 'no cortar' — la impresión añadirá corte automáticamente)

                    def do_print():
                        # Intentar imprimir directamente a una impresora POS en Windows usando win32print (pywin32).
                        # Si no está disponible o falla, caer al flujo anterior que genera un PDF y lo envía a imprimir.
                        try:
                            import sys, tempfile, os
                            # Solo intentamos escritura directa en Windows (POS normalmente conectado/instalado allí)
                            if sys.platform.startswith('win'):
                                try:
                                    import win32print
                                    # Resolver impresora seleccionada o predeterminada
                                    try:
                                        from app_config import get_printer_name
                                        sel = get_printer_name()
                                        printer_name = sel if sel else win32print.GetDefaultPrinter()
                                    except Exception:
                                        printer_name = win32print.GetDefaultPrinter()
                                    hPrinter = win32print.OpenPrinter(printer_name)
                                    try:
                                        # Preparar texto del ticket como bytes; al final añadiremos
                                        # secuencias ESC/POS para forzar corte del papel.
                                        text = "\n".join(ticket) + "\n\n"
                                        # Probar codificaciones comunes para impresoras POS
                                        try:
                                            data = text.encode('cp437', errors='replace')
                                        except Exception:
                                            try:
                                                data = text.encode('cp1252', errors='replace')
                                            except Exception:
                                                data = text.encode('utf-8', errors='replace')

                                        # Añadir secuencias comunes de corte de papel (ESC/POS)
                                        try:
                                            # GS V 0 (partial cut) or GS V 1 (full) — algunos drivers usan 0
                                            cut_seqs = [b'\x1dV\x00', b'\x1dV\x01', b'\x1b\x69', b'\x1b\x6d']
                                            # añadir nuevas líneas y luego la primera secuencia válida
                                            data = data + b'\n\n\n'
                                            # concatenar una secuencia de corte (no todas para evitar duplicados extremos)
                                            data = data + cut_seqs[0]
                                        except Exception:
                                            pass

                                        # Enviar como RAW al spooler
                                        win32print.StartDocPrinter(hPrinter, 1, ("Ticket", None, "RAW"))
                                        win32print.StartPagePrinter(hPrinter)
                                        win32print.WritePrinter(hPrinter, data)
                                        win32print.EndPagePrinter(hPrinter)
                                        win32print.EndDocPrinter(hPrinter)
                                        messagebox.showinfo('Imprimir', f'Enviado a impresora: {printer_name}')
                                        return
                                    finally:
                                        try:
                                            win32print.ClosePrinter(hPrinter)
                                        except Exception:
                                            pass
                                except ImportError:
                                    # pywin32 no disponible: caemos a PDF
                                    pass
                                except Exception as e:
                                    # Error al usar win32print: mostrar aviso y caer a PDF
                                    messagebox.showwarning('Imprimir', f'No se pudo imprimir directamente: {e}\nSe intentará por PDF.')

                            # Fallback: generar un PDF temporal y enviarlo a imprimir (método seguro)
                            from reportlab.lib.pagesizes import A4
                            from reportlab.pdfgen import canvas
                            fd, pdf_path = tempfile.mkstemp(suffix='.pdf')
                            os.close(fd)
                            c = canvas.Canvas(pdf_path, pagesize=A4)
                            y = 800
                            c.setFont('Helvetica', 10)
                            for line in ticket:
                                c.drawString(40, y, str(line))
                                y -= 14
                                if y < 60:
                                    c.showPage()
                                    y = 800
                            c.save()
                            if sys.platform.startswith('win'):
                                os.startfile(pdf_path, 'print')
                            else:
                                import subprocess
                                subprocess.Popen(['lp', pdf_path])
                        except Exception as e:
                            messagebox.showerror('Imprimir', f'Error al imprimir: {e}')

                    def do_export_pdf():
                        # Exportar a PDF: usar reportlab si está disponible
                        try:
                            from reportlab.lib.pagesizes import A4
                            from reportlab.pdfgen import canvas
                            import tempfile, os
                            fd, path = tempfile.mkstemp(suffix='.pdf')
                            os.close(fd)
                            c = canvas.Canvas(path, pagesize=A4)
                            y = 800
                            c.setFont('Helvetica', 10)
                            for line in ticket:
                                c.drawString(40, y, str(line))
                                y -= 14
                                if y < 60:
                                    c.showPage()
                                    y = 800
                            c.save()
                            # abrir el pdf generado
                            if os.name == 'nt':
                                os.startfile(path)
                            else:
                                import subprocess
                                subprocess.Popen(['xdg-open', path])
                        except Exception:
                            # fallback: guardar como .txt
                            try:
                                from tkinter import filedialog as _fd
                                fpath = _fd.asksaveasfilename(defaultextension='.txt', filetypes=[('Text','*.txt')])
                                if fpath:
                                    with open(fpath, 'w', encoding='utf-8') as f:
                                        f.write(preview)
                                    messagebox.showinfo('Exportar', f'Archivo guardado: {fpath}')
                            except Exception as e:
                                messagebox.showerror('Exportar', f'Error exportando: {e}')

                    tk.Button(btn_frame, text='Imprimir', command=do_print).pack(side='left', padx=6)
                    tk.Button(btn_frame, text='Exportar a PDF', command=do_export_pdf).pack(side='left', padx=6)
                except Exception:
                    # Fallback simple si no se puede abrir Toplevel
                    messagebox.showinfo('Vista previa del ticket', preview)
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al imprimir: {str(e)}")