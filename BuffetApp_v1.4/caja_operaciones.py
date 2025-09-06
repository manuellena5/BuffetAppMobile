# archivo: caja_operaciones.py
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import datetime
from db_utils import get_connection
import sqlite3
from theme import (
    COLORS, FONTS, FINANCE_COLORS, format_currency,
    themed_button, apply_theme
)
class DetalleCajaFrame(tk.Frame):
    def __init__(self, parent, caja_id, on_close=None, disable_movimientos=True):
        super().__init__(parent)
        self.caja_id = caja_id
        self.on_close = on_close
        # Flag to indicate we want to disable ingreso/retiro while this frame is open
        self.disable_movimientos = disable_movimientos
        self.config(bg=COLORS['background'], padx=15, pady=15)
        
        # Grid configuration
        self.columnconfigure(0, weight=3)  # KPIs y detalles
        self.columnconfigure(1, weight=2)  # Panel de cierre
        
        self._create_kpi_panel()
        self._create_tables_panel()
        self._create_closure_panel()
        self._load_data()
        # Ask parent (usually main app) to disable menu movimientos while detail is open
        try:
            if self.disable_movimientos and hasattr(self.master, 'actualizar_menu_caja'):
                # set caja info so menu state will disable ingreso/retiro
                # call a helper if exists to force disable
                if hasattr(self.master, 'menu_bar') and hasattr(self.master, 'caja_menu'):
                    # directly disable the menu entries if possible
                    try:
                        self.master.caja_menu.entryconfig("Ingreso de efectivo", state='disabled')
                        self.master.caja_menu.entryconfig("Retiro de efectivo", state='disabled')
                    except Exception:
                        pass
        except Exception:
            pass
        
    def _create_kpi_panel(self):
        # Panel principal de KPIs
        kpi_frame = tk.Frame(self, bg=COLORS['background'])
        kpi_frame.grid(row=0, column=0, sticky="nsew", pady=(0,15))
        
        # Frame para los KPIs
        self.kpis = tk.Frame(kpi_frame, bg=COLORS['background'])
        self.kpis.pack(fill='x', pady=(0,10))
        
    def create_kpi(self, parent, icon, title, value, bg_color, fg_color):
        f = tk.Frame(parent, bg=bg_color, padx=10, pady=5)
        f.pack(side='left', padx=5)
        tk.Label(f, text=f"{icon} {title}", bg=bg_color, fg=fg_color, 
                font=FONTS['normal']).pack()
        tk.Label(f, text=value, bg=bg_color, fg=fg_color,
                font=FONTS['bold']).pack()
        return f
        
    def _create_tables_panel(self):
        tables_frame = tk.Frame(self, bg=COLORS['background'])
        tables_frame.grid(row=1, column=0, sticky="nsew")
        
        # Ventas por categoría
        cat_frame = tk.LabelFrame(tables_frame, text="🗂 Ventas por categoría",
                                bg=COLORS['surface'], font=FONTS['bold'])
        cat_frame.pack(fill='x', pady=(0,10))
        
        self.cat_tree = ttk.Treeview(cat_frame, columns=('categoria', 'monto'),
                                   show='headings', height=6)
        self.cat_tree.heading('categoria', text='Categoría')
        self.cat_tree.heading('monto', text='Monto')
        self.cat_tree.pack(fill='x', padx=5, pady=5)
        
        # Productos vendidos
        prod_frame = tk.LabelFrame(tables_frame, text="📦 Productos vendidos",
                                 bg=COLORS['surface'], font=FONTS['bold'])
        prod_frame.pack(fill='x')
        
        self.prod_tree = ttk.Treeview(prod_frame, 
                                    columns=('producto', 'cant', 'monto'),
                                    show='headings', height=8)
        self.prod_tree.heading('producto', text='Producto')
        self.prod_tree.heading('cant', text='Cant')
        self.prod_tree.heading('monto', text='Monto')
        self.prod_tree.column('cant', width=70, anchor='center')
        self.prod_tree.column('monto', width=100, anchor='e')
        self.prod_tree.pack(fill='x', padx=5, pady=5)
        
    def _create_closure_panel(self):
        closure_frame = tk.LabelFrame(self, text="Cierre de Caja",
                                      bg=COLORS['surface'], font=FONTS['bold'])
        closure_frame.grid(row=0, column=1, rowspan=2, sticky="nsew", padx=(15,0))
        # Observaciones de apertura (solo lectura)
        tk.Label(closure_frame, text="📝 Observaciones de apertura:",
                 bg=COLORS['surface']).pack(anchor='w', pady=(10,0))
        self.obs_apertura_text = tk.Text(closure_frame, height=2, font=FONTS['normal'], state='disabled', bg=COLORS.get('disabled_bg', '#f5f5f5'), fg=COLORS.get('disabled_fg', '#666'))
        self.obs_apertura_text.pack(fill='x', pady=(0,10))
        # Observaciones de movimientos (solo lectura)
        tk.Label(closure_frame, text="📝 Observaciones de movimientos:",
                 bg=COLORS['surface']).pack(anchor='w')
        self.obs_movimientos_text = tk.Text(closure_frame, height=2, font=FONTS['normal'], state='disabled', bg=COLORS.get('disabled_bg', '#f5f5f5'), fg=COLORS.get('disabled_fg', '#666'))
        self.obs_movimientos_text.pack(fill='x', pady=(0,10))
        # Campos de cierre (editables si la caja está abierta)
        self.campos_cierre_frame = tk.Frame(closure_frame, bg=COLORS['surface'])
        self.campos_cierre_frame.pack(fill='x', pady=5)
        # Conteo efectivo
        tk.Label(self.campos_cierre_frame, text="💵 Conteo efectivo final en caja:",
                 bg=COLORS['surface']).pack(anchor='w')
        self.conteo_entry = tk.Entry(self.campos_cierre_frame, font=FONTS['normal'], disabledbackground=COLORS.get('disabled_bg'), disabledforeground=COLORS.get('disabled_fg'))
        self.conteo_entry.pack(fill='x', pady=(0,10))
        # Transferencias
        tk.Label(self.campos_cierre_frame, text="🔁 Monto transferencias:",
                 bg=COLORS['surface']).pack(anchor='w')
        self.transf_entry = tk.Entry(self.campos_cierre_frame, font=FONTS['normal'], disabledbackground=COLORS.get('disabled_bg'), disabledforeground=COLORS.get('disabled_fg'))
        self.transf_entry.pack(fill='x', pady=(0,10))
        # Usuario cierre
        tk.Label(self.campos_cierre_frame, text="👤 Usuario cierre:",
                 bg=COLORS['surface']).pack(anchor='w')
        self.usuario_entry = tk.Entry(self.campos_cierre_frame, font=FONTS['normal'], disabledbackground=COLORS.get('disabled_bg'), disabledforeground=COLORS.get('disabled_fg'))
        self.usuario_entry.pack(fill='x', pady=(0,10))
        # Observaciones de cierre
        tk.Label(self.campos_cierre_frame, text="📝 Observaciones de cierre:",
                 bg=COLORS['surface']).pack(anchor='w')
        self.obs_text = tk.Text(self.campos_cierre_frame, height=3, font=FONTS['normal'])
        self.obs_text.pack(fill='x', pady=(0,10))
        # Diferencia calculada
        self.diff_label = tk.Label(closure_frame, text="🧮 Diferencia: $ 0,00",
                                   font=FONTS['bold'], bg=COLORS['surface'])
        self.diff_label.pack(pady=10)
        # Botones
        btn_frame = tk.Frame(closure_frame, bg=COLORS['surface'])
        btn_frame.pack(fill='x', pady=10)
        self.btn_cerrar = themed_button(btn_frame, text="⚠️ Cerrar Caja", command=self._cerrar_caja,
                                        bg='#F44336', fg='white')
        self.btn_cerrar.pack(fill='x', pady=(0,5))
        self.btn_imprimir = themed_button(btn_frame, text="🖨️ Imprimir Ticket",
                                          command=self._imprimir_ticket)
        self.btn_imprimir.pack(fill='x', pady=2)
        themed_button(btn_frame, text="📊 Exportar Excel",
                     command=self._exportar_excel).pack(fill='x', pady=2)
        themed_button(btn_frame, text="📄 Exportar PDF",
                     command=self._exportar_pdf).pack(fill='x', pady=2)
        themed_button(btn_frame, text="❌ Salir",
                     command=self._salir).pack(fill='x', pady=(10,0))
        # Bind para cálculo en vivo
        self.conteo_entry.bind('<KeyRelease>', self._calcular_diferencia)
        self.transf_entry.bind('<KeyRelease>', self._calcular_diferencia)
    def _salir(self):
        """Cierra la vista de detalle"""
        if self.on_close:
            # Before closing, re-enable movimiento menu items on the main app if possible
            try:
                if self.disable_movimientos and hasattr(self.master, 'caja_menu'):
                    try:
                        if hasattr(self.master, 'actualizar_menu_caja'):
                            self.master.actualizar_menu_caja()
                        else:
                            self.master.caja_menu.entryconfig("Ingreso de efectivo", state='normal')
                            self.master.caja_menu.entryconfig("Retiro de efectivo", state='normal')
                    except Exception:
                        pass
            except Exception:
                pass
            # Report that the detail was closed but the caja was NOT closed
            try:
                self.on_close(False)
            except TypeError:
                # backward compatibility: allow handlers without parameter
                self.on_close()
            
    def _calcular_diferencia(self, event=None):
        try:
            conteo = float(self.conteo_entry.get().replace(',', '.') or 0)
            transf = float(self.transf_entry.get().replace(',', '.') or 0)
            # Según fórmula requerida:
            # Teórico = Fondo inicial + Ventas totales teóricas + Ingresos - Retiros
            # Real    = Conteo efectivo + Transferencias + Ingresos - Retiros + Fondo inicial
            ingresos = float(getattr(self, 'ingresos', 0) or 0)
            retiros = float(getattr(self, 'retiros', 0) or 0)
            fondo = float(getattr(self, 'fondo_inicial', 0) or 0)
            total_real = conteo + transf + ingresos - retiros + fondo
            teorico = float(getattr(self, 'total_teorico', 0) or 0)
            diferencia = total_real - teorico
            
            color = FINANCE_COLORS['positive_fg'] if diferencia >= 0 else FINANCE_COLORS['negative_fg']
            self.diff_label.config(
                text=f"🧮 Diferencia: {format_currency(diferencia, include_sign=True)}",
                fg=color
            )
        except ValueError:
            self.diff_label.config(
                text="🧮 Diferencia: $ 0,00", 
                fg=COLORS['text']
            )
            
    def _load_data(self):
        with get_connection() as conn:
            cursor = conn.cursor()
            # Cargar datos básicos de la caja (seleccionar columnas explícitas para evitar dependencias de índice)
            try:
                cursor.execute("""
                    SELECT cd.id, cd.codigo_caja, cd.fecha, cd.hora_apertura, cd.hora_cierre,
                           cd.usuario_apertura, cd.usuario_cierre, cd.fondo_inicial,
                           cd.observaciones_apertura, cd.obs_cierre, cd.total_ventas,
                           cd.total_efectivo_teorico, cd.conteo_efectivo_final, cd.transferencias_final,
                           cd.ingresos, cd.retiros, cd.diferencia, cd.total_tickets,
                           d.descripcion as nombre_disciplina
                    FROM caja_diaria cd
                    LEFT JOIN disciplinas d ON d.codigo = cd.disciplina
                    WHERE cd.id = ?
                """, (self.caja_id,))
                caja = cursor.fetchone()
            except sqlite3.Error:
                caja = None
            if not caja:
                messagebox.showerror("Error", "No se encontró la caja especificada")
                if self.on_close:
                    self.on_close()
                return
            # Calcular totales (seguro ante tablas faltantes)
            total_ventas = 0
            try:
                cursor.execute("""
                    SELECT COALESCE(SUM(t.total_ticket),0) as total
                    FROM tickets t 
                    JOIN ventas v ON v.id = t.venta_id
                    WHERE v.caja_id = ? AND t.status != 'Anulado'
                """, (self.caja_id,))
                total_ventas = cursor.fetchone()[0] or 0
            except sqlite3.Error:
                total_ventas = 0
            # Mapear campos de caja a variables legibles (si la consulta devolvió datos)
            try:
                # Indices según SELECT anterior - guardar como atributos de instancia para usarlos en export
                self.codigo_caja = caja[1]
                self.fecha = caja[2]
                self.hora_apertura = caja[3]
                self.hora_cierre = caja[4]
                self.usuario_apertura = caja[5]
                self.usuario_cierre = caja[6]
                self.fondo_inicial = caja[7] or 0
                self.observaciones_apertura = caja[8] or ''
                self.obs_cierre_db = caja[9] or ''
                # total_ventas puede venir vacío en la tabla
                self.total_ventas = caja[10] or total_ventas or 0
                try:
                    self.total_teorico = caja[11] if caja[11] is not None else None
                except Exception:
                    self.total_teorico = None
                self.conteo_efectivo_final = caja[12] or 0
                self.transferencias_final = caja[13] or 0
                self.ingresos = caja[14] or 0
                self.retiros = caja[15] or 0
                self.diferencia_db = caja[16]
                self.total_tickets = caja[17] or 0
                # contar tickets anulados explícitamente
                try:
                    cursor.execute("SELECT COALESCE(COUNT(*),0) FROM tickets t JOIN ventas v ON v.id = t.venta_id WHERE v.caja_id=? AND t.status='Anulado'", (self.caja_id,))
                    self.tickets_anulados = cursor.fetchone()[0] or 0
                except Exception:
                    self.tickets_anulados = 0
                self.nombre_disciplina = caja[18] or ''
            except Exception:
                # Si algo falla, inicializar valores por seguridad
                self.codigo_caja = None
                self.fecha = self.hora_apertura = self.hora_cierre = self.usuario_apertura = self.usuario_cierre = ''
                self.fondo_inicial = 0
                self.observaciones_apertura = self.obs_cierre_db = ''
                self.total_ventas = total_ventas or 0
                self.total_teorico = getattr(self, 'total_teorico', 0)
                self.conteo_efectivo_final = self.transferencias_final = self.ingresos = self.retiros = 0
                self.diferencia_db = None
                self.total_tickets = 0
                self.nombre_disciplina = ''
            # Si total_teorico no está presente en la fila, estimar con una suma segura
            if getattr(self, 'total_teorico', None) is None:
                try:
                    cursor.execute("""
                        SELECT COALESCE(SUM(t.total_ticket),0)
                        FROM tickets t
                        JOIN ventas v ON v.id = t.venta_id
                        LEFT JOIN metodos_pago mp ON mp.id = v.metodo_pago_id
                        WHERE v.caja_id = ? AND (mp.descripcion='Efectivo' OR mp.descripcion IS NULL) AND t.status != 'Anulado'
                    """, (self.caja_id,))
                    efectivo_ventas = cursor.fetchone()[0] or 0
                except Exception:
                    efectivo_ventas = total_ventas
                # Estimación: fondo inicial + efectivo ventas + ingresos - retiros + transferencias
                self.total_teorico = self.fondo_inicial + efectivo_ventas + self.ingresos - self.retiros + self.transferencias_final
            # Poblamos observaciones de apertura
            try:
                self.obs_apertura_text.config(state='normal')
                self.obs_apertura_text.delete('1.0', tk.END)
                self.obs_apertura_text.insert('1.0', getattr(self, 'observaciones_apertura', ''))
                self.obs_apertura_text.config(state='disabled')
            except Exception:
                pass
            # Listar movimientos y sus observaciones en el textbox correspondiente
            movimientos_text = []
            try:
                cursor.execute("SELECT tipo, monto, observacion, creado_ts FROM caja_movimiento WHERE caja_id=? ORDER BY creado_ts", (self.caja_id,))
                for tipo, monto, observacion, creado in cursor.fetchall():
                    movimientos_text.append(f"{tipo}: {monto} - {observacion or ''} ({creado})")
            except Exception:
                movimientos_text = []
            try:
                self.obs_movimientos_text.config(state='normal')
                self.obs_movimientos_text.delete('1.0', tk.END)
                if movimientos_text:
                    self.obs_movimientos_text.insert('1.0', '\n'.join(movimientos_text))
                self.obs_movimientos_text.config(state='disabled')
            except Exception:
                pass
            # Colocar observacion de cierre en el campo editable
            try:
                self.obs_text.delete('1.0', tk.END)
                if getattr(self, 'obs_cierre_db', ''):
                    self.obs_text.insert('1.0', getattr(self, 'obs_cierre_db', ''))
            except Exception:
                pass
            # Rellenar los campos de cierre con los valores guardados (si existen)
            try:
                # conteo, transferencias, usuario
                try:
                    self.conteo_entry.delete(0, tk.END)
                    self.conteo_entry.insert(0, str(getattr(self, 'conteo_efectivo_final', '') or ''))
                except Exception:
                    pass
                try:
                    self.transf_entry.delete(0, tk.END)
                    self.transf_entry.insert(0, str(getattr(self, 'transferencias_final', '') or ''))
                except Exception:
                    pass
                try:
                    self.usuario_entry.delete(0, tk.END)
                    self.usuario_entry.insert(0, str(getattr(self, 'usuario_cierre', '') or ''))
                except Exception:
                    pass
            except Exception:
                pass
            # Ventas por método de pago
            try:
                cursor.execute("""
                    SELECT mp.descripcion, COUNT(*) as cantidad, SUM(t.total_ticket) as total
                    FROM ventas v
                    JOIN tickets t ON t.venta_id = v.id
                    LEFT JOIN metodos_pago mp ON mp.id = v.metodo_pago_id
                    WHERE v.caja_id = ? AND t.status != 'Anulado'
                    GROUP BY mp.descripcion
                """, (self.caja_id,))
                ventas_por_metodo = cursor.fetchall()
            except sqlite3.Error:
                ventas_por_metodo = []
            # KPIs con valores calculados y consistentes
            try:
                # Fondo inicial
                self.create_kpi(
                    self.kpis, "💰", "Fondo Inicial",
                    format_currency(getattr(self, 'fondo_inicial', 0)),
                    FINANCE_COLORS.get('transfer_bg', '#eee'),
                    FINANCE_COLORS.get('transfer_fg', '#000')
                )
            except Exception:
                self.create_kpi(self.kpis, "💰", "Fondo Inicial", format_currency(getattr(self, 'fondo_inicial', 0)), '#ffffff', '#000000')
            # Obtener ingresos y retiros preferentemente desde caja_movimiento si no vienen en la fila
            try:
                ingresos_val = getattr(self, 'ingresos', None)
                retiros_val = getattr(self, 'retiros', None)
                if ingresos_val is None:
                    cursor.execute("SELECT COALESCE(SUM(monto),0) FROM caja_movimiento WHERE caja_id = ? AND tipo = 'INGRESO'", (self.caja_id,))
                    ingresos_val = cursor.fetchone()[0] or 0
                if retiros_val is None:
                    cursor.execute("SELECT COALESCE(SUM(monto),0) FROM caja_movimiento WHERE caja_id = ? AND tipo = 'RETIRO'", (self.caja_id,))
                    retiros_val = cursor.fetchone()[0] or 0
            except Exception:
                ingresos_val = getattr(self, 'ingresos', 0) or 0
                retiros_val = getattr(self, 'retiros', 0) or 0
            try:
                self.create_kpi(self.kpis, "⬆️", "Ingresos", format_currency(ingresos_val), FINANCE_COLORS.get('positive_bg', '#e6ffed'), FINANCE_COLORS.get('positive_fg', '#0a0'))
                self.create_kpi(self.kpis, "⬇️", "Retiros", format_currency(retiros_val), FINANCE_COLORS.get('negative_bg', '#ffecec'), FINANCE_COLORS.get('negative_fg', '#c00'))
            except Exception:
                self.create_kpi(self.kpis, "⬆️", "Ingresos", format_currency(ingresos_val), '#ffffff', '#000000')
                self.create_kpi(self.kpis, "⬇️", "Retiros", format_currency(retiros_val), '#ffffff', '#000000')
            # Ventas por medio de pago sumas (efectivo / transferencias)
            total_efectivo = sum(total for metodo, cant, total in ventas_por_metodo if (metodo or '').lower() == 'efectivo') if ventas_por_metodo else 0
            total_transf = sum(total for metodo, cant, total in ventas_por_metodo if (metodo or '').lower() != 'efectivo') if ventas_por_metodo else 0
            try:
                self.create_kpi(self.kpis, "💵", "Efectivo", format_currency(total_efectivo), FINANCE_COLORS.get('transfer_bg', '#fff'), FINANCE_COLORS.get('transfer_fg', '#000'))
                self.create_kpi(self.kpis, "🔁", "Transferencias", format_currency(total_transf), FINANCE_COLORS.get('transfer_bg', '#fff'), FINANCE_COLORS.get('transfer_fg', '#000'))
            except Exception:
                self.create_kpi(self.kpis, "💵", "Efectivo", format_currency(total_efectivo), '#ffffff', '#000000')
                self.create_kpi(self.kpis, "🔁", "Transferencias", format_currency(total_transf), '#ffffff', '#000000')
            # KPI: tickets anulados
            try:
                self.create_kpi(self.kpis, "🚫", "Anulados", str(getattr(self, 'tickets_anulados', 0)), COLORS.get('surface'), COLORS.get('text'))
            except Exception:
                try:
                    self.create_kpi(self.kpis, "🚫", "Anulados", str(getattr(self, 'tickets_anulados', 0)), '#ffffff', '#000')
                except Exception:
                    pass
            # Ventas totales (destacado)
            try:
                ventas_frame = tk.Frame(self.kpis, bg=FINANCE_COLORS.get('total_sales_bg', COLORS.get('surface', '#f5f5f5')))
                ventas_frame.pack(side='left', padx=5, fill='y')
                tk.Label(ventas_frame, text="📊 Ventas totales", bg=FINANCE_COLORS.get('total_sales_bg', COLORS.get('surface', '#f5f5f5')), fg=FINANCE_COLORS.get('total_sales_fg', COLORS.get('text', '#000')), font=FONTS.get('title')).pack(padx=10, pady=(5,0))
                tk.Label(ventas_frame, text=format_currency(getattr(self, 'total_ventas', total_ventas)), bg=FINANCE_COLORS.get('total_sales_bg', COLORS.get('surface', '#f5f5f5')), fg=FINANCE_COLORS.get('total_sales_fg', COLORS.get('text', '#000')), font=FONTS.get('title')).pack(padx=10, pady=(0,5))
            except Exception:
                pass
            # Cargar tablas: categorías
            try:
                cursor.execute("""
                    SELECT c.descripcion, SUM(t.total_ticket) as total
                    FROM tickets t
                    JOIN ventas v ON v.id = t.venta_id
                    LEFT JOIN Categoria_Producto c ON c.id = t.categoria_id
                    WHERE v.caja_id = ? AND t.status != 'Anulado'
                    GROUP BY c.descripcion
                    ORDER BY total DESC
                """, (self.caja_id,))
                rows = cursor.fetchall()
            except sqlite3.Error:
                rows = []
            for cat in self.cat_tree.get_children():
                self.cat_tree.delete(cat)
            for row in rows:
                cat = row[0] or 'Sin categoría'
                total = row[1] or 0
                self.cat_tree.insert('', 'end', values=(cat, format_currency(total)))
            # Productos vendidos
            try:
                cursor.execute("""
                    SELECT p.nombre, SUM(vi.cantidad) as cant, 
                           SUM(vi.cantidad * vi.precio_unitario) as total
                    FROM venta_items vi
                    JOIN tickets t ON t.id = vi.ticket_id
                    JOIN ventas v ON v.id = t.venta_id
                    JOIN products p ON p.id = vi.producto_id
                    WHERE v.caja_id = ? AND t.status != 'Anulado'
                    GROUP BY p.nombre
                    ORDER BY total DESC
                """, (self.caja_id,))
                prod_rows = cursor.fetchall()
            except sqlite3.Error:
                prod_rows = []
            for prod in self.prod_tree.get_children():
                self.prod_tree.delete(prod)
            for row in prod_rows:
                try:
                    self.prod_tree.insert('', 'end', values=(row[0], int(row[1]), format_currency(row[2] or 0)))
                except Exception:
                    pass
            # Obtener estado actual (cerrada/abierta) desde la DB (usar la fila ya leída)
            try:
                # En la consulta inicial, la columna 'estado' no fue seleccionada; leer desde tabla
                cursor.execute("SELECT estado FROM caja_diaria WHERE id=?", (self.caja_id,))
                row = cursor.fetchone()
                esta_cerrada = (row and row[0] == 'cerrada')
            except Exception:
                esta_cerrada = False
            # Ajustar estado de campos de cierre según si la caja está cerrada
            if esta_cerrada:
                # Deshabilitar campos editables del cierre
                try:
                    self.conteo_entry.config(state='disabled')
                    self.transf_entry.config(state='disabled')
                    self.usuario_entry.config(state='disabled')
                    # Observaciones de cierre no editables (gris)
                    self.obs_text.config(state='disabled', bg=COLORS.get('disabled_bg', '#f5f5f5'))
                except Exception:
                    pass
                # Ocultar o deshabilitar botón de cerrar y mostrar/habilitar imprimir
                try:
                    # preferimos deshabilitar botón de cerrar si existe
                    self.btn_cerrar.config(state='disabled')
                except Exception:
                    try:
                        self.btn_cerrar.pack_forget()
                    except Exception:
                        pass
                try:
                    self.btn_imprimir.config(state='normal')
                    self.btn_imprimir.pack(fill='x', pady=2)
                except Exception:
                    try:
                        self.btn_imprimir.pack(fill='x', pady=2)
                    except Exception:
                        pass
            else:
                # Caja abierta: campos editables y botón de cerrar activo
                try:
                    self.conteo_entry.config(state='normal')
                    self.transf_entry.config(state='normal')
                    self.usuario_entry.config(state='normal')
                    self.obs_text.config(state='normal')
                except Exception:
                    pass
                try:
                    self.btn_cerrar.config(state='normal')
                    self.btn_cerrar.pack(fill='x', pady=(0,5))
                except Exception:
                    try:
                        self.btn_cerrar.pack(fill='x', pady=(0,5))
                    except Exception:
                        pass
                try:
                    # ocultar o desactivar imprimir si la caja no está cerrada
                    self.btn_imprimir.config(state='disabled')
                    self.btn_imprimir.pack_forget()
                except Exception:
                    try:
                        self.btn_imprimir.pack_forget()
                    except Exception:
                        pass
            # Guardar total teórico (usar columna total_efectivo_teorico si existe)
            try:
                self.total_teorico = caja[14] or total_ventas
            except Exception:
                self.total_teorico = total_ventas
            
    def _cerrar_caja(self):
        try:
            conteo = float(self.conteo_entry.get().replace(',', '.') or 0)
            transf = float(self.transf_entry.get().replace(',', '.') or 0)
            usuario = self.usuario_entry.get().strip()
            obs = self.obs_text.get("1.0", tk.END).strip()
            
            if not usuario:
                messagebox.showwarning("Validación", "Debe ingresar el usuario de cierre")
                return
                
            if messagebox.askyesno("Confirmar", 
                                 "¿Está seguro de cerrar la caja? Esta acción no se puede deshacer."):
                with get_connection() as conn:
                    cursor = conn.cursor()
                    now = datetime.datetime.now()
                    # Calcular diferencia con la fórmula completa
                    ingresos = float(getattr(self, 'ingresos', 0) or 0)
                    retiros = float(getattr(self, 'retiros', 0) or 0)
                    fondo = float(getattr(self, 'fondo_inicial', 0) or 0)
                    teorico = float(getattr(self, 'total_teorico', 0) or 0)
                    real = conteo + transf + ingresos - retiros + fondo
                    diferencia = real - teorico
                    cursor.execute("""
                        UPDATE caja_diaria 
                        SET estado = 'cerrada',
                            hora_cierre = ?,
                            usuario_cierre = ?,
                            conteo_efectivo_final = ?,
                            transferencias_final = ?,
                            diferencia = ?,
                            obs_cierre = ?
                        WHERE id = ?
                    """, (
                        now.strftime('%H:%M:%S'),
                        usuario,
                        conteo,
                        transf,
                        diferencia,
                        obs,
                        self.caja_id
                    ))
                    conn.commit()
                    messagebox.showinfo("Éxito", "Caja cerrada correctamente")
                    # Notify caller that the caja was actually closed
                    if self.on_close:
                        try:
                            self.on_close(True)
                        except TypeError:
                            # backward compatibility
                            self.on_close()
        
        except ValueError:
            messagebox.showerror("Error", 
                               "Los montos deben ser números válidos")
    
    def _exportar_excel(self):
        try:
            # Nombre por defecto usando el codigo legible de la caja si está disponible
            default_code = getattr(self, 'codigo_caja', None) or self.caja_id
            default_name = f"{default_code}_CierreCaja.xlsx"
            filename = filedialog.asksaveasfilename(
                initialfile=default_name,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Guardar reporte de cierre de caja"
            )
            if not filename:
                return
            # Intentar usar openpyxl si está disponible
            try:
                import openpyxl
                from openpyxl.utils import get_column_letter
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "CierreCaja"
                # Cabecera con columnas solicitadas
                columns = [
                    'Codigo Caja', 'Fecha', 'Hora Apertura', 'Hora Cierre',
                    'Usuario Apertura', 'Usuario Cierre', 'Disciplina',
                    'Fondo inicial', 'Total ventas', 'Total efectivo teorico',
                    'Conteo efectivo final', 'Transferencias final', 'Ingresos', 'Retiros',
                    'Diferencia', 'Total tickets', 'Observacion apertura', 'Observacion cierre',
                    'Movimientos', 'Items vendidos'
                ]
                ws.append(columns)
                # Preparar valores (leer desde atributos / widgets con fallbacks)
                try:
                    movimientos_joined = self.obs_movimientos_text.get('1.0', tk.END).strip().replace('\n', ' | ')
                except Exception:
                    movimientos_joined = ''
                # Construir cadena con items vendidos: Producto (Cant x subtotal)
                try:
                    with get_connection() as conn2:
                        cur2 = conn2.cursor()
                        cur2.execute("""
                            SELECT p.nombre, SUM(vi.cantidad) as cant, SUM(vi.cantidad * vi.precio_unitario) as total
                            FROM venta_items vi
                            JOIN tickets t ON t.id = vi.ticket_id
                            JOIN ventas v ON v.id = t.venta_id
                            JOIN products p ON p.id = vi.producto_id
                            WHERE v.caja_id = ?
                            GROUP BY p.nombre
                        """, (self.caja_id,))
                        items_rows = cur2.fetchall()
                    items_joined = ' | '.join(f"{r[0]} ({int(r[1])} x {format_currency(r[2])})" for r in items_rows)
                except Exception:
                    items_joined = ''
                row = [
                    getattr(self, 'codigo_caja', self.caja_id),
                    getattr(self, 'fecha', ''),
                    getattr(self, 'hora_apertura', ''),
                    getattr(self, 'hora_cierre', ''),
                    getattr(self, 'usuario_apertura', ''),
                    getattr(self, 'usuario_cierre', ''),
                    getattr(self, 'nombre_disciplina', ''),
                    getattr(self, 'fondo_inicial', self.conteo_entry.get()),
                    getattr(self, 'total_ventas', ''),
                    getattr(self, 'total_teorico', ''),
                    getattr(self, 'conteo_efectivo_final', self.conteo_entry.get()),
                    getattr(self, 'transferencias_final', self.transf_entry.get()),
                    getattr(self, 'ingresos', ''),
                    getattr(self, 'retiros', ''),
                    getattr(self, 'diferencia_db', self.diff_label.cget('text')),
                    getattr(self, 'total_tickets', ''),
                    getattr(self, 'observaciones_apertura', ''),
                    getattr(self, 'obs_cierre_db', ''),
                    movimientos_joined,
                    items_joined
                ]
                ws.append(row)
                # Ajustar anchos
                for i in range(1, len(columns) + 1):
                    ws.column_dimensions[get_column_letter(i)].width = 25
                wb.save(filename)
            except Exception:
                # fallback simple CSV-like TXT
                with open(filename, 'w', encoding='utf-8') as f:
                    cols = ['Codigo Caja','Fecha','Hora Apertura','Hora Cierre','Usuario Apertura','Usuario Cierre','Disciplina','Fondo inicial','Total ventas','Total efectivo teorico','Conteo efectivo final','Transferencias final','Ingresos','Retiros','Diferencia','Total tickets','Observacion apertura','Observacion cierre','Movimientos','Items vendidos']
                    f.write(';'.join(cols) + '\n')
                    try:
                        mov = self.obs_movimientos_text.get('1.0', tk.END).strip().replace('\n', ' | ')
                    except Exception:
                        mov = ''
                    try:
                        items_joined_val = items_joined
                    except Exception:
                        items_joined_val = ''
                    values = [
                        str(getattr(self, 'codigo_caja', self.caja_id)),
                        str(getattr(self, 'fecha', '')),
                        str(getattr(self, 'hora_apertura', '')),
                        str(getattr(self, 'hora_cierre', '')),
                        str(getattr(self, 'usuario_apertura', '')),
                        str(getattr(self, 'usuario_cierre', '')),
                        str(getattr(self, 'nombre_disciplina', '')),
                        str(getattr(self, 'fondo_inicial', self.conteo_entry.get())),
                        str(getattr(self, 'total_ventas', '')),
                        str(getattr(self, 'total_teorico', '')),
                        str(getattr(self, 'conteo_efectivo_final', self.conteo_entry.get())),
                        str(getattr(self, 'transferencias_final', self.transf_entry.get())),
                        str(getattr(self, 'ingresos', '')),
                        str(getattr(self, 'retiros', '')),
                        str(getattr(self, 'diferencia_db', self.diff_label.cget('text'))),
                        str(getattr(self, 'total_tickets', '')),
                        str(getattr(self, 'observaciones_apertura', '')),
                        str(getattr(self, 'obs_cierre_db', '')),
                        str(mov),
                        str(items_joined_val)
                    ]
                    f.write(';'.join(values) + '\n')
            # Abrir archivo automáticamente
            try:
                import os, subprocess, sys
                if sys.platform.startswith('win'):
                    os.startfile(filename)
                else:
                    subprocess.Popen(['xdg-open', filename])
            except Exception:
                pass
            messagebox.showinfo("Éxito", f"Archivo guardado: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {str(e)}")
    
    def _exportar_pdf(self):
        try:
            default_code = getattr(self, 'codigo_caja', None) or self.caja_id
            default_name = f"{default_code}_CierreCaja.pdf"
            filename = filedialog.asksaveasfilename(
                initialfile=default_name,
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Guardar reporte de cierre de caja"
            )
            if not filename:
                return
            # Intentar usar reportlab
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas
                c = canvas.Canvas(filename, pagesize=A4)
                y = 800
                c.setFont("Helvetica-Bold", 14)
                c.drawCentredString(300, y, "CIERRE DE CAJA")
                y -= 40
                c.setFont("Helvetica", 10)
                c.drawString(50, y, f"Codigo Caja: {getattr(self, 'codigo_caja', self.caja_id)}")
                y -= 20
                c.drawString(50, y, f"Fondo inicial: {getattr(self, 'fondo_inicial', self.conteo_entry.get())}")
                y -= 20
                c.drawString(50, y, f"Transferencias: {getattr(self, 'transferencias_final', self.transf_entry.get())}")
                y -= 20
                c.drawString(50, y, f"Diferencia: {self.diff_label.cget('text')}")
                y -= 20
                try:
                    c.drawString(50, y, f"Tickets anulados: {getattr(self, 'tickets_anulados', 0)}")
                    y -= 20
                except Exception:
                    pass
                # incluir items vendidos
                try:
                    with get_connection() as conn3:
                        ccur = conn3.cursor()
                        ccur.execute("""
                            SELECT p.nombre, SUM(vi.cantidad) as cant, SUM(vi.cantidad * vi.precio_unitario) as total
                            FROM venta_items vi
                            JOIN tickets t ON t.id = vi.ticket_id
                            JOIN ventas v ON v.id = t.venta_id
                            JOIN products p ON p.id = vi.producto_id
                            WHERE v.caja_id = ?
                            GROUP BY p.nombre
                        """, (self.caja_id,))
                        items_rows_pdf = ccur.fetchall()
                    if items_rows_pdf:
                        y -= 10
                        c.setFont("Helvetica-Bold", 12)
                        c.drawString(50, y, "Items vendidos:")
                        y -= 18
                        c.setFont("Helvetica", 10)
                        for r in items_rows_pdf:
                            line = f"{r[0]} ({int(r[1])} x {format_currency(r[2])})"
                            c.drawString(60, y, line)
                            y -= 14
                            if y < 60:
                                c.showPage()
                                y = 800
                except Exception:
                    pass
                c.save()
            except Exception:
                # fallback: write plain text
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write('CIERRE DE CAJA\n')
                    f.write(f"Codigo Caja: {getattr(self, 'codigo_caja', self.caja_id)}\n")
                    f.write(f"Fondo inicial: {getattr(self, 'fondo_inicial', self.conteo_entry.get())}\n")
                    f.write(f"Transferencias: {getattr(self, 'transferencias_final', self.transf_entry.get())}\n")
                    f.write(f"Diferencia: {self.diff_label.cget('text')}\n")
                    try:
                        f.write(f"Tickets anulados: {getattr(self, 'tickets_anulados', 0)}\n")
                    except Exception:
                        pass
                    # incluir items vendidos en texto
                    try:
                        with get_connection() as conn4:
                            tcur = conn4.cursor()
                            tcur.execute("""
                                SELECT p.nombre, SUM(vi.cantidad) as cant, SUM(vi.cantidad * vi.precio_unitario) as total
                                FROM venta_items vi
                                JOIN tickets t ON t.id = vi.ticket_id
                                JOIN ventas v ON v.id = t.venta_id
                                JOIN products p ON p.id = vi.producto_id
                                WHERE v.caja_id = ?
                                GROUP BY p.nombre
                            """, (self.caja_id,))
                            trows = tcur.fetchall()
                            if trows:
                                f.write('\nItems vendidos:\n')
                                for r in trows:
                                    f.write(f"- {r[0]} ({int(r[1])} x {format_currency(r[2])})\n")
                    except Exception:
                        pass
            # abrir automaticamente
            try:
                import os, sys, subprocess
                if sys.platform.startswith('win'):
                    os.startfile(filename)
                else:
                    subprocess.Popen(['xdg-open', filename])
            except Exception:
                pass
            messagebox.showinfo("Éxito", f"Archivo guardado: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {str(e)}")
            
    def _imprimir_ticket(self):
        """Imprime el ticket de cierre de caja"""
        try:
            with get_connection() as conn:
                cursor = conn.cursor()
                
                # Obtener datos de la caja
                cursor.execute("""
                    SELECT cd.*, d.descripcion as disciplina
                    FROM caja_diaria cd
                    LEFT JOIN disciplinas d ON d.codigo = cd.disciplina
                    WHERE cd.id = ?
                """, (self.caja_id,))
                caja = cursor.fetchone()
                
                if not caja:
                    messagebox.showerror("Error", "No se encontró la caja")
                    return
                
                # Componer el ticket
                ticket = []
                ticket.append("=" * 40)
                ticket.append("CIERRE DE CAJA".center(40))
                ticket.append("=" * 40)
                ticket.append(f"Fecha: {caja[1]}")
                ticket.append(f"Apertura: {caja[2]}")
                ticket.append(f"Cierre: {caja[3]}")
                ticket.append(f"Responsable: {caja[5]}")
                ticket.append(f"Disciplina: {caja[-1]}")
                ticket.append("-" * 40)
                
                # Totales por método de pago
                cursor.execute("""
                    SELECT mp.descripcion, SUM(t.total_ticket) as total
                    FROM ventas v
                    JOIN tickets t ON t.venta_id = v.id
                    LEFT JOIN metodos_pago mp ON mp.id = v.metodo_pago_id
                    WHERE v.caja_id = ? AND t.status != 'Anulado'
                    GROUP BY mp.descripcion
                """, (self.caja_id,))
                
                ticket.append("TOTALES POR MEDIO DE PAGO")
                ticket.append("-" * 40)
                total_general = 0
                for metodo, total in cursor.fetchall():
                    ticket.append(f"{metodo}: {format_currency(total)}")
                    total_general += total
                
                ticket.append("-" * 40)
                ticket.append(f"TOTAL: {format_currency(total_general)}")
                ticket.append("-" * 40)
                
                # Información de cierre
                ticket.append(f"Fondo inicial: {format_currency(caja[8])}")
                ticket.append(f"Conteo final: {format_currency(caja[10])}")
                ticket.append(f"Transferencias: {format_currency(caja[11])}")
                ticket.append(f"Diferencia: {format_currency(caja[12])}")
                # intentar añadir anulados si la columna existe/está presente
                try:
                    ticket.append(f"Tickets anulados: {getattr(self, 'tickets_anulados', '0')}")
                except Exception:
                    pass
                ticket.append("=" * 40)
                
                # TODO: Implementar impresión física
                # Por ahora solo mostrar vista previa
                preview = "\n".join(ticket)
                messagebox.showinfo("Vista previa del ticket", preview)
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al imprimir: {str(e)}")